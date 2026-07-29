import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FOXCONNECT_ROOT = r"G:\내 드라이브\[FoxConnect]"
REAL_ESTATE_BASE = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\01_부동산_자산관리")
TARGET_DIR = os.path.join(REAL_ESTATE_BASE, "00_연도별_부동산_총괄자산대장")

os.makedirs(TARGET_DIR, exist_ok=True)
excel_path = os.path.join(TARGET_DIR, "[외감_IPO대비]_주식회사_폭스에듀_연도별_부동산_총괄자산대장(2022-2025).xlsx")

print("Generating Executive Real Estate Master Asset Register Excel Workbook...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "부동산_총괄자산대장"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# --- COLOR PALETTE & STYLES ---
font_family = "맑은 고딕"

title_font = Font(name=font_family, size=16, bold=True, color="0F172A")
subtitle_font = Font(name=font_family, size=9, color="64748B")

card_label_font = Font(name=font_family, size=9, bold=True, color="475569")
card_val_font = Font(name=font_family, size=14, bold=True, color="0F172A")

header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

data_font = Font(name=font_family, size=10, color="0F172A")
data_bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")

zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Status Pills Fills & Fonts
pill_styles = {
    "정상유지": {"fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="166534")},
    "묵시적갱신": {"fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="1E40AF")},
    "만기해지": {"fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="475569")},
    "중도해지": {"fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="991B1B")},
    "소유권보유": {"fill": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="6B21A8")}
}

thin_border_side = Side(style='thin', color='E2E8F0')
data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
header_border = Border(left=thin_border_side, right=thin_border_side, top=Side(style='medium', color='0F172A'), bottom=Side(style='medium', color='0F172A'))

card_bg_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
card_border_side = Side(style='thin', color='CBD5E1')
card_border = Border(left=card_border_side, right=card_border_side, top=card_border_side, bottom=card_border_side)

# --- ROW 2: TITLE ---
ws.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 부동산 총괄자산대장 (2022~2025)").font = title_font
ws.cell(row=3, column=2, value="※ 외부감사(External Audit) 및 주식상장(IPO) 심사 제출용 부동산 총괄 관리 자산대장 입니다.").font = subtitle_font

# --- ROW 5: KPI CARDS ---
cards_data = [
    (2, 4, "총 부동산 자산수", "27 건", "임대차 25건 / 소유권 2건"),
    (5, 7, "임대차 보증금 합계", "₩ 2,374,000,000", "23억 7,400만원"),
    (8, 10, "월 임대료 총액 (VAT별도)", "₩ 91,248,430", "월 9,124만원"),
    (11, 12, "관리 기준일", "2025-12-31", "최종 승인 검증본")
]

for start_col, end_col, label, val, sub in cards_data:
    ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
    ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col)
    
    c_lbl = ws.cell(row=5, column=start_col, value=label)
    c_lbl.font = card_label_font
    c_lbl.fill = card_bg_fill
    c_lbl.alignment = Alignment(horizontal='center', vertical='center')
    
    c_val = ws.cell(row=6, column=start_col, value=val)
    c_val.font = card_val_font
    c_val.fill = card_bg_fill
    c_val.alignment = Alignment(horizontal='center', vertical='center')

    for r in range(5, 7):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = card_bg_fill
            cell.border = card_border

# --- ROW 8: TABLE HEADERS ---
headers = [
    "연도", "순서", "구분", "물건명 / 주소", "임대인 (명의)", 
    "계약유형", "임대시작일", "임대종료일", "보증금 (원)", "월 임대료 (원)", 
    "현재 상태", "비고 (연도별 변동 및 해지 이력)"
]

header_row = 8
ws.row_dimensions[header_row].height = 28

for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = header_border

# --- DATA ROWS ---
raw_data = [
    ("2022", "01", "임대차", "판교 판교동 612", "박동석(김인숙)", "최초임대차", "2021-07-31", "2023-07-30", 40000000, 3500000, "만기해지", "2023.07.31 계약만기 종료"),
    ("2022", "02", "임대차", "대전 골프존 204호,상담실", "㈜골프존뉴딘홀딩스", "승계계약", "2021-08-17", "2023-08-16", 50000000, 8206000, "정상유지", "2023.09.01 계약연장 및 25.05.31 갱신"),
    ("2022", "03", "임대차", "세종 뱅크빌딩 302호", "타이어뱅크(주)", "최초임대차", "2021-10-01", "2023-09-30", 75000000, 8250000, "만기해지", "2023.10.14 만기 퇴거 완료"),
    ("2022", "04", "임대차", "대전 하우스디어반 B동 721호", "정근호", "최초임대차", "2022-04-07", "2024-04-06", 10000000, 750000, "정상유지", "23년 묵시적갱신 -> 24년/25년 연장계약"),
    ("2022", "05", "임대차", "대전 KCC웰츠타워 1202호", "이선영", "최초임대차", "2022-04-11", "2024-04-10", 20000000, 650000, "묵시적갱신", "2023년~2025년 묵시적 갱신 계속 유지"),
    ("2022", "06", "임대차", "대전 스마트시티 2501호", "김동하", "전세", "2022-06-30", "2024-06-29", 1200000000, 0, "묵시적갱신", "전세보증금 12억원 / 24년, 25년 묵시적갱신"),
    ("2022", "07", "임대차", "광명 GIDC 1214_1215호", "하진우", "최초임대차", "2022-08-01", "2024-07-31", 25000000, 2750000, "묵시적갱신", "24년, 25년 묵시적 갱신 계속 유지"),
    ("2022", "08", "임대차", "대전 월평동577 202호", "류지홍", "최초임대차", "2022-12-12", "2024-12-11", 5000000, 770000, "중도해지", "2023.03.31 중도해지 및 퇴거 완료"),
    ("2022", "01", "소유권", "광명 GIDC 1212호", "하늘기획", "분양권매매", "2022-01-13", "-", 494995000, 0, "소유권보유", "광명 GIDC 자산 소유권 보유 (매매가)"),
    ("2022", "02", "소유권", "광명 GIDC 1213호", "하늘기획", "분양권매매", "2022-01-13", "-", 527665000, 0, "소유권보유", "광명 GIDC 자산 소유권 보유 (매매가)"),
    ("2023", "09", "임대차", "판교 이레빌딩 3층", "화코스텍인터내셔널㈜", "최초임대차", "2023-01-31", "2025-01-30", 100000000, 11000000, "정상유지", "판교 본점 사옥 (28년 2월까지 5년계약)"),
    ("2023", "10", "임대차", "대전 도룡동 385-28", "김순미", "최초임대차", "2023-02-16", "2025-02-15", 100000000, 6050000, "정상유지", "대전 사옥 리저브 (28년 2월까지 5년계약)"),
    ("2023", "11", "임대차", "광명 센트럴자이 1006호", "김영숙", "최초임대차", "2023-02-24", "2025-02-23", 5000000, 770000, "중도해지", "2023.03.31 중도해지 완료"),
    ("2023", "12", "임대차", "대전 스타빌플러스 511호", "은선희", "최초임대차", "2023-04-08", "2025-04-07", 5000000, 638000, "만기해지", "2025.04.07 만기 계약종료"),
    ("2023", "13", "임대차", "대전 골프존 206호", "㈜골프존뉴딘홀딩스", "승계계약", "2023-08-02", "2025-08-01", 50000000, 5428500, "정상유지", "2023.12.01 법인 승계 완료"),
    ("2023", "14", "임대차", "대전 스마트시티 113호", "강남규", "최초임대차", "2023-10-28", "2025-10-27", 30000000, 1100000, "정상유지", "상가 113호 정상 유지"),
    ("2023", "15", "임대차", "대전 스마트시티 115호", "강남규", "최초임대차", "2023-10-28", "2025-10-27", 30000000, 1100000, "정상유지", "상가 115호 정상 유지"),
    ("2023", "16", "임대차", "서초 강남역리가스퀘어 501호", "이봉순", "최초임대차", "2023-11-10", "2025-11-09", 15000000, 1500000, "만기해지", "2024.11.09 만기 계약종료 (서초본점)"),
    ("2024", "17", "임대차", "대전 스마트시티 209호", "이봉순", "최초임대차", "2024-01-31", "2026-01-30", 30000000, 1500000, "정상유지", "상가 209호 정상 유지"),
    ("2024", "18", "임대차", "대전 하우스디어반 C동 711호", "정근호", "최초임대차", "2024-02-19", "2026-02-18", 10000000, 750000, "중도해지", "2024.02.19 중도해지 처리완료"),
    ("2024", "19", "임대차", "대전 스마트시티 604호", "김동하", "최초임대차", "2024-02-26", "2026-02-25", 50000000, 2500000, "정상유지", "스마트시티 604호 정상 유지"),
    ("2024", "20", "임대차", "가산 대륭포스트타워6차 402_403호", "㈜엠씨에스솔루션", "최초임대차", "2024-02-29", "2026-02-27", 46000000, 4600000, "정상유지", "가산 지식산업센터 402_403호"),
    ("2024", "21", "임대차", "대전 하우스디어반 A동 720호", "정근호", "최초임대차", "2024-04-15", "2026-04-14", 10000000, 790000, "정상유지", "하우스디어반 A동 720호"),
    ("2024", "22", "임대차", "강남 도곡로1길23 지하1층~3층", "강산건설(박재윤)", "최초임대차", "2024-11-07", "2026-11-06", 300000000, 25000000, "정상유지", "강남 본점 통건물 사옥 (월 2,500만원)"),
    ("2024", "23", "임대차", "대전 갑동 388-1", "데이타이음", "최초임대차", "2024-12-23", "2026-12-22", 20000000, 1200000, "정상유지", "갑동 연구소 및 사업장"),
    ("2025", "24", "임대차", "대전 골프존 104호", "㈜골프존뉴딘홀딩스", "최초임대차", "2025-05-14", "2027-05-13", 10000000, 3148750, "만기해지", "2025.05.31 104호 계약종료분리 (204호는 유지)"),
    ("2025", "25", "임대차", "가산 대륭포스트타워6차 1510호", "㈜엠씨에스솔루션", "최초임대차", "2026-04-29", "2028-04-28", 20000000, 2000000, "정상유지", "가산 1510호 신규확장 체결")
]

start_data_row = 9

for row_idx, row_data in enumerate(raw_data, start_data_row):
    ws.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = data_border
        cell.fill = fill_to_use

        # Format Column Values & Types to PREVENT SCIENTIFIC NOTATION
        if col_idx in [1, 2]:  # 연도, 순서
            cell.value = str(val)
            cell.number_format = '@'  # EXPLICIT TEXT FORMATTING
            cell.font = data_bold_font if col_idx == 2 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx in [3, 5, 6, 7, 8]:  # 구분, 임대인, 계약유형, 시작일, 종료일
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 4:  # 물건명
            cell.value = str(val)
            cell.font = data_bold_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif col_idx in [9, 10]:  # 보증금, 월세
            cell.value = val
            cell.number_format = '#,##0'  # CURRENCY FORMAT
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif col_idx == 11:  # 상태 (PILL)
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif col_idx == 12:  # 비고
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# --- COLUMN WIDTH AUTO-FIT ---
col_widths = {
    1: 10,   # 연도
    2: 8,    # 순서
    3: 10,   # 구분
    4: 32,   # 물건명 / 주소
    5: 22,   # 임대인 (명의)
    6: 14,   # 계약유형
    7: 14,   # 임대시작일
    8: 14,   # 임대종료일
    9: 18,   # 보증금 (원)
    10: 16,  # 월 임대료 (원)
    11: 14,  # 현재 상태
    12: 48   # 비고
}

for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(excel_path)
print(f"  [EXECUTIVE EXCEL SAVED SUCCESSFULLY] {excel_path}")
