import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

FOXCONNECT_ROOT = r"G:\내 드라이브\[FoxConnect]"
veh_excel_p = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\02_차량_자산관리\00_연도별_차량_총괄자산대장\[외감_IPO대비]_주식회사_폭스에듀_연도별_법인차량_총괄자산대장(2022-2025).xlsx")

print("Inspecting Vehicle Excel File:", veh_excel_p)

if os.path.exists(veh_excel_p):
    wb = openpyxl.load_workbook(veh_excel_p)
    ws = wb["01_법인차량_자산대장"]
    
    print("\n--- Headers (Row 7) ---")
    headers = [ws.cell(row=7, column=c).value for c in range(2, 14)]
    print(headers)
    
    print("\n--- Data Rows (Row 8 to 17) ---")
    for r in range(8, 18):
        row_vals = [ws.cell(row=r, column=c).value for c in range(2, 14)]
        print(f"Row {r:2d}:", row_vals)
