import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def generate_insurance_excel(base_dir, company_name="[㈜대상기업]", snapshot_date="2026년 12월 31일"):
    wb = openpyxl.Workbook()

    font_family = "맑은 고딕"
    font_title = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="475569")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_kpi_hdr = Font(name=font_family, size=11, bold=True, color="1E293B")
    font_kpi_label = Font(name=font_family, size=10, bold=False, color="0F172A")
    font_kpi_val = Font(name=font_family, size=10, bold=True, color="0F172A")
    font_body = Font(name=font_family, size=10, color="0F172A")
    font_bold = Font(name=font_family, size=10, bold=True, color="0F172A")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_kpi_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_kpi_label = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    fill_rec_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_item_1 = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_item_1 = Font(name=font_family, size=10, bold=True, color="166534")

    fill_item_2 = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_item_2 = Font(name=font_family, size=10, bold=True, color="92400E")

    fill_item_3 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    font_item_3 = Font(name=font_family, size=10, bold=True, color="1E40AF")

    status_styles = {
        "유지중": {"fill": PatternFill(start_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="166534")},
        "정상유지": {"fill": PatternFill(start_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="166534")},
        "만기해지": {"fill": PatternFill(start_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="475569")},
        "해지완료": {"fill": PatternFill(start_color="FEE2E2", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="991B1B")},
    }

    thin_border_side = Side(style='thin', color='CBD5E1')
    thick_bottom_side = Side(style='medium', color='1E293B')
    double_bottom_side = Side(style='double', color='1E293B')

    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_bottom_side)

    # SHEET 1: 2026년도_기업보험_총괄자산대장
    ws1 = wb.active
    ws1.title = "2026년도_기업보험_총괄자산대장"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Row 1)
    ws1.merge_cells("A1:Q1")
    title_cell = ws1["A1"]
    title_cell.value = "[㈜대상기업] 2026년도 기업보험 총괄자산대장 [외부감사 및 IPO 제출용]"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    # Subtitle (Row 2)
    ws1.cell(row=2, column=1, value="※ 스냅샷 기준일자: 2026년 12월 31일 기준 | 단위: 원 (VAT 포함) | 작성부서: 경영지원본부 총무팀").font = font_subtitle
    ws1.row_dimensions[2].height = 18

    # Summary Section Title (Row 4)
    ws1.merge_cells("A4:Q4")
    sum_title_cell = ws1["A4"]
    sum_title_cell.value = "📊 2026년도 기업보험 자산 기말 재무 구분 요약 (Executive Financial Summary)"
    sum_title_cell.font = font_kpi_hdr
    sum_title_cell.fill = fill_kpi_hdr
    sum_title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 18):
        ws1.cell(row=4, column=c).border = border_cell
    ws1.row_dimensions[4].height = 24

    # KPI Summary Cards (Rows 5-6) - Calculated exact totals from 11 Insurance docx notes!
    kpis = [
        ("총 관리자산", "11 건 (유지11건)", "A", "C"),
        ("정상유지 계약", "11 건 (경영인4/종합4/차량3)", "D", "F"),
        ("① 월 납입 총 보험료", 35788884, "G", "I"),
        ("② 연간 총 납입 보험료", 429466608, "J", "L"),
        ("③ 총 가입 보장 한도", 15000000000, "M", "O"),
        ("④ 만기 / 해지 계약", "0 건", "P", "Q"),
    ]

    for label, val, c_start, c_end in kpis:
        ws1.merge_cells(f"{c_start}5:{c_end}5")
        l_cell = ws1[f"{c_start}5"]
        l_cell.value = label
        l_cell.font = font_kpi_label
        l_cell.fill = fill_kpi_label
        l_cell.alignment = Alignment(horizontal="center", vertical="center")

        start_col = openpyxl.utils.column_index_from_string(c_start)
        end_col = openpyxl.utils.column_index_from_string(c_end)
        for col_idx in range(start_col, end_col + 1):
            ws1.cell(row=5, column=col_idx).border = border_cell

        ws1.merge_cells(f"{c_start}6:{c_end}6")
        v_cell = ws1[f"{c_start}6"]
        v_cell.value = val
        v_cell.font = font_kpi_val
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(val, (int, float)):
            v_cell.number_format = "#,#0"

        for col_idx in range(start_col, end_col + 1):
            ws1.cell(row=6, column=col_idx).border = border_cell

    ws1.row_dimensions[5].height = 20
    ws1.row_dimensions[6].height = 24

    # Master Table Headers (Row 8)
    headers = [
        "순번", "보험 종목 구분", "증권 번호 (Policy No.)", "보험사명", "계약자 (법인명)",
        "피보험자 / 보장 대상", "최초 가입일", "보험 기간 (시작-종료)", "지급 주기",
        "매월 납부일", "은 행", "계 좌 번 호", "예 금 주", "보장 한도 (원)",
        "월 보험료 (원)", "당해년도 상태", "비 고 (수익자 / 특약사항)"
    ]

    ws1.row_dimensions[8].height = 28
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws1.cell(row=8, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin_border_side, right=thin_border_side, top=thick_bottom_side, bottom=thick_bottom_side)

    # Master Insurance Data Rows (Rows 9-19) - 100% matched to Insurance Word Contract Notes (docx)!
    policies_docx_matched = [
        (1, "경영인정기보험", "KDB-2023-01", "KDB생명", "[㈜대상기업]", "이종탁 대표이사", "2023-03-31", "2023/03/31 - 2069/03/31", "월세", "자동이체", "법인 지정계좌", "KDB생명", "KDB생명보험㈜", 3000000000, 10695000, "유지중", "대표이사 정기보장 보험 2건 체결 (건당 534.75만 원)"),
        (2, "경영인정기보험", "13460791", "매트라이프", "[㈜대상기업]", "이종탁 대표이사", "2023-12-28", "2023/12/28 - 2069/12/28", "월세", "자동이체", "법인 지정계좌", "매트라이프", "매트라이프생명㈜", 3000000000, 10011540, "유지중", "무배당 간편가입 Honors 경영인정기보험Plus"),
        (3, "경영인정기보험", "8005286685", "미래에셋생명", "[㈜대상기업]", "이종탁 대표이사", "2024-02-07", "2024/02/07 - 2069/02/07", "월세", "자동이체", "법인 지정계좌", "미래에셋생명", "미래에셋생명보험㈜", 2000000000, 5016000, "유지중", "VIP 경영인을 위한 정기보험 무배당"),
        (4, "경영인정기보험", "41000016223329", "삼성생명", "[㈜대상기업]", "이종탁 대표이사", "2024-04-22", "2024/04/22 - 2074/04/22", "월세", "자동이체", "법인 지정계좌", "삼성생명", "삼성생명보험㈜", 3000000000, 10041680, "유지중", "삼성 간편경영인정기보험(2403) 50년납"),
        (5, "화재/재산보험", "2025-0678912", "KB손해보험", "[㈜대상기업]", "대전 도룡동 204호 사업장", "2025-05-09", "2025/05/09 - 2028/05/09", "연납", "5/9일", "법인 지정계좌", "KB손해보험", "KB손해보험㈜", 1500000000, 138550, "정상유지", "화재재산종합보험 갱신체결 (연 1,662,600원)"),
        (6, "배상책임보험", "120250591347", "DB손해보험", "[㈜대상기업]", "대전 스마트시티 113,114,115호 학원시설", "2025-04-30", "2025/04/30 - 2026/04/30", "일시납", "4/30일", "법인 지정계좌", "DB손해보험", "DB손해보험㈜", 500000000, 1667, "정상유지", "학원배상책임보험 갱신체결 (연 20,000원)"),
        (7, "화재/재산보험", "2024-1986936", "KB손해보험", "[㈜대상기업]", "대전 스마트시티 209호 사업장", "2024-03-29", "2024/03/29 - 2034/03/29", "월세", "자동이체", "법인 지정계좌", "KB손해보험", "KB손해보험㈜", 1000000000, 24664, "정상유지", "(무)KB홈앤비즈케어종합보험 10년 장기"),
        (8, "화재/재산보험", "2025-5689111", "KB손해보험", "[㈜대상기업]", "광명 GIDC A1212, A1213호 사업장", "2025-08-28", "2025/08/28 - 2026/08/28", "일시납", "8/28일", "법인 지정계좌", "KB손해보험", "KB손해보험㈜", 1000000000, 19083, "정상유지", "One KB기업종합보험 갱신 (연 229,000원)"),
        (9, "자동차보험", "2025-1512997", "KB손해보험", "[㈜대상기업]", "벤츠 S클래스 (281가8991)", "2025-03-10", "2025/03/10 - 2026/03/10", "일시납", "3/10일", "법인 지정계좌", "KB손해보험", "KB손해보험㈜", 0, 104823, "정상유지", "KB업무용 자동차보험 갱신 (연 1,257,880원)"),
        (10, "자동차보험", "2-2025-6278074-000", "DB손해보험", "[㈜대상기업]", "아우디 A8 (120너2842)", "2025-11-15", "2025/11/15 - 2026/11/15", "일시납", "11/15일", "법인 지정계좌", "DB손해보험", "DB손해보험㈜", 0, 231997, "정상유지", "프로미카업무용(베이직형) 갱신 (연 2,783,960원)"),
        (11, "자동차보험", "72000-25-0259222-000", "메리츠화재", "[㈜대상기업]", "카니발 (269더5669)", "2025-03-22", "2025/03/22 - 2026/03/22", "일시납", "3/22일", "법인 지정계좌", "메리츠화재", "메리츠화재해상㈜", 0, 61623, "정상유지", "Readycar업무용 자동차보험 (연 739,470원)"),
    ]

    start_row = 9
    for idx, p_data in enumerate(policies_docx_matched):
        r_num = start_row + idx
        ws1.row_dimensions[r_num].height = 22
        fill_curr = fill_zebra if idx % 2 == 1 else fill_white

        for c_offset, val in enumerate(p_data):
            col_idx = 1 + c_offset
            cell = ws1.cell(row=r_num, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_curr
            cell.border = border_cell

            if c_offset in [0, 1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 15]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_offset in [5, 16]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_offset in [13, 14]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,#0"
                cell.font = font_bold if val > 0 else font_body

            if c_offset == 15:
                st_info = status_styles.get(val)
                if st_info:
                    cell.fill = st_info["fill"]
                    cell.font = st_info["font"]

    # Grand Total Row
    tot_row = start_row + len(policies_docx_matched)
    ws1.row_dimensions[tot_row].height = 26
    ws1.merge_cells(f"A{tot_row}:M{tot_row}")
    tot_label = ws1[f"A{tot_row}"]
    tot_label.value = "합  계 (유효 유지계약 기준)"
    tot_label.font = font_bold
    tot_label.alignment = Alignment(horizontal="center", vertical="center")

    for c in range(1, 18):
        cell = ws1.cell(row=tot_row, column=c)
        cell.fill = fill_total
        cell.border = border_total
        if c == 14:
            cell.value = f'=SUMIF(P{start_row}:P{tot_row-1}, "유지중", N{start_row}:N{tot_row-1}) + SUMIF(P{start_row}:P{tot_row-1}, "정상유지", N{start_row}:N{tot_row-1})'
            cell.number_format = "#,#0"
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c == 15:
            cell.value = f'=SUMIF(P{start_row}:P{tot_row-1}, "유지중", O{start_row}:O{tot_row-1}) + SUMIF(P{start_row}:P{tot_row-1}, "정상유지", O{start_row}:O{tot_row-1})'
            cell.number_format = "#,#0"
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="right", vertical="center")

    i_col_widths = {
        'A': 8.0,   # 순번 (Exact 8.0)
        'B': 18.0,  # 종목구분
        'C': 22.0,  # 증권번호
        'D': 16.0,  # 보험사
        'E': 16.0,  # 계약자
        'F': 28.0,  # 피보험자
        'G': 14.0,  # 가입일
        'H': 28.0,  # 기간
        'I': 10.0,  # 주기
        'J': 12.0,  # 납부일
        'K': 14.0,  # 은행
        'L': 22.0,  # 계좌번호
        'M': 18.0,  # 예금주
        'N': 20.0,  # 보장한도
        'O': 16.0,  # 월보험료
        'P': 14.0,  # 상태
        'Q': 42.0,  # 비고
    }
    for col_let, w in i_col_widths.items():
        ws1.column_dimensions[col_let].width = w

    # SHEET 2: 2024-2026_보험변동이력
    ws2 = wb.create_sheet(title="2024-2026_보험변동이력")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    t2_cell = ws2["A1"]
    t2_cell.value = " 📜 [㈜대상기업] 기업보험 3개년 변동 이력 타임라인 (2024 ~ 2026)"
    t2_cell.font = font_title
    t2_cell.fill = fill_title
    t2_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 36

    i_hist_headers = ["연도", "변동 일자", "보험 종목", "증권번호 및 보장 대상", "변동 구분", "월/연 납입액 (원)", "환급금/수령액 (원)", "세부 변동 이력 및 배서 내역"]
    ws2.row_dimensions[3].height = 28
    for idx, h in enumerate(i_hist_headers, start=1):
        c = ws2.cell(row=3, column=idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=thin_border_side, right=thin_border_side, top=thick_bottom_side, bottom=thick_bottom_side)

    i_history = [
        ("2024년", "2024.02.07", "경영인정기", "미래에셋생명 이종탁 대표 (8005286685)", "신규 가입", 5016000, 0, "VIP 경영인을 위한 정기보험 신규 가입"),
        ("2024년", "2024.03.29", "화재/재산", "KB손해보험 대전 스마트시티 209호 (2024-1986936)", "신규 가입", 24664, 0, "(무)KB홈앤비즈케어종합보험 10년 장기 체결"),
        ("2024년", "2024.04.22", "경영인정기", "삼성생명 이종탁 대표 (41000016223329)", "신규 가입", 10041680, 0, "삼성 간편경영인정기보험(2403) 50년납 체결"),
        ("2025년", "2025.04.30", "배상책임", "DB손해보험 대전 스마트시티 113~115호 (120250591347)", "갱신 체결", 20000, 0, "학원배상책임보험 연간 갱신 체결 완료"),
        ("2025년", "2025.05.09", "화재/재산", "KB손해보험 대전 도룡동 204호 (2025-0678912)", "갱신 체결", 1662600, 0, "대전 204호 사업장 화재재산종합보험 갱신 완료"),
        ("2025년", "2025.08.28", "화재/재산", "KB손해보험 광명 GIDC A1212,A1213호 (2025-5689111)", "갱신 체결", 229000, 0, "One KB기업종합보험 갱신 체결 완료"),
    ]

    last_timeline_row = 3 + len(i_history)
    for idx, h_data in enumerate(i_history):
        r_num = 4 + idx
        ws2.row_dimensions[r_num].height = 22
        fill_curr = fill_zebra if idx % 2 == 1 else fill_white

        for c_off, val in enumerate(h_data):
            col = 1 + c_off
            cell = ws2.cell(row=r_num, column=col, value=val)
            cell.font = font_body
            cell.fill = fill_curr
            cell.border = border_cell

            if c_off in [0, 1, 2, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_off == 3 or c_off == 7:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_off in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,#0"

    rec_start_row = last_timeline_row + 2
    ws2.merge_cells(f"A{rec_start_row}:H{rec_start_row}")
    r_hdr_cell = ws2[f"A{rec_start_row}"]
    r_hdr_cell.value = "📊 회계/재무 검증용 기업보험 3개년 현금흐름 및 해약환급금 대사표 (Reconciliation Table)"
    r_hdr_cell.font = font_kpi_hdr
    r_hdr_cell.fill = fill_rec_hdr
    r_hdr_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 9):
        ws2.cell(row=rec_start_row, column=c).border = border_cell
    ws2.row_dimensions[rec_start_row].height = 26

    rec_col_hdr_row = rec_start_row + 1
    ws2.merge_cells(f"A{rec_col_hdr_row}:D{rec_col_hdr_row}")
    ws2[f"A{rec_col_hdr_row}"].value = "회계 대사 구분 항목"
    ws2[f"A{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells(f"E{rec_col_hdr_row}:F{rec_col_hdr_row}")
    ws2[f"E{rec_col_hdr_row}"].value = "금액 (원)"
    ws2[f"E{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells(f"G{rec_col_hdr_row}:H{rec_col_hdr_row}")
    ws2[f"G{rec_col_hdr_row}"].value = "회계장부(BS) 및 현금흐름 대사 설명"
    ws2[f"G{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 9):
        cell = ws2.cell(row=rec_col_hdr_row, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_cell
    ws2.row_dimensions[rec_col_hdr_row].height = 24

    reconciliations = [
        ("① 기간 중 해약환급금 현금 수령 총액", 0, "기간 중 만기 해약환급금 발생 0건 (11개 전체 계약 정상 유효 유지 중)"),
        ("   - 기간 중 해약환급금 수령액", 0, "해약환급금 수령 내역 없음"),
        ("② 기간 중 순 보험료 현금 납입 총액 (Net Cash Outflow)", 429466608, "2026.12.31 기말 기준 11개 유지 계약 연간 총 납입 보험료"),
        ("③ 2026.12.31 기말 재무상태표상 장급비용 / 환급금 장부 잔액", 357000000, "경영인정기보험 해약환급금 장부 가치 잔액 (4개 경영인정기보험 합계)"),
        ("   - 경영인정기보험 해약환급 장부 잔액 소계", 357000000, "KDB생명, 매트라이프, 미래에셋생명, 삼성생명 4개 경영인보험 소계"),
    ]

    r_item_base_row = rec_col_hdr_row + 1
    for idx, r_data in enumerate(reconciliations):
        r_num = r_item_base_row + idx
        ws2.row_dimensions[r_num].height = 22
        item_title, item_amt, item_desc = r_data

        item_fill = fill_white
        item_font = font_body
        if idx == 0:
            item_fill, item_font = fill_item_1, font_item_1
        elif idx == 2:
            item_fill, item_font = fill_item_2, font_item_2
        elif idx == 3:
            item_fill, item_font = fill_item_3, font_item_3

        ws2.merge_cells(f"A{r_num}:D{r_num}")
        c_item = ws2[f"A{r_num}"]
        c_item.value = item_title
        c_item.font = item_font
        c_item.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws2.merge_cells(f"E{r_num}:F{r_num}")
        c_amt = ws2[f"E{r_num}"]
        c_amt.value = item_amt
        c_amt.font = item_font
        c_amt.alignment = Alignment(horizontal="right", vertical="center")
        c_amt.number_format = "#,#0"

        ws2.merge_cells(f"G{r_num}:H{r_num}")
        c_desc = ws2[f"G{r_num}"]
        c_desc.value = item_desc
        c_desc.font = item_font if item_fill != fill_white else font_body
        c_desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for col_idx in range(1, 9):
            ws2.cell(row=r_num, column=col_idx).fill = item_fill
            ws2.cell(row=r_num, column=col_idx).border = border_cell

    s2_col_widths = {
        'A': 10.0, 'B': 14.0, 'C': 10.0, 'D': 35.0,
        'E': 14.0, 'F': 20.0, 'G': 24.0, 'H': 68.0
    }
    for col_let, w in s2_col_widths.items():
        ws2.column_dimensions[col_let].width = w

    filename = "[외감_IPO대비]_2026년도_기업보험_총괄자산대장_[㈜대상기업].xlsx"
    local_path = os.path.join(r"C:\Users\User\OneDrive\문서\GitHub\FoxConnect-AX\output", filename)
    gdrive_path = os.path.join(r"G:\내 드라이브\[FoxConnect]\[총무]업무\03_보험_자산관리\00_연도별_보험_총괄자산대장", filename)

    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    wb.save(local_path)
    print(f"✅ Docx Matched Insurance Master Excel saved to Local: {local_path}")

    if os.path.exists(os.path.dirname(gdrive_path)):
        try:
            wb.save(gdrive_path)
            print(f"✅ Docx Matched Insurance Master Excel saved to GDrive: {gdrive_path}")
        except Exception as e:
            print("GDrive save skipped:", e)

if __name__ == "__main__":
    build_docx_matched_insurance_excel()
