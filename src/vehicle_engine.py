import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def generate_vehicle_excel(base_dir, company_name="[㈜대상기업]", snapshot_date="2026년 12월 31일"):
    wb = openpyxl.Workbook()

    font_family = "맑은 고딕"
    font_title = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="475569")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_kpi_hdr = Font(name=font_family, size=11, bold=True, color="1E293B")
    font_kpi_label = Font(name=font_family, size=10, bold=False, color="0F172A")
    font_kpi_val = Font(name=font_family, size=10, bold=True, color="0F172A")
    font_body = Font(name=font_family, size=10, color="0F172A")
    font_bold = Font(name=font_family, size=10, bold=True, color="0F172A")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_kpi_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_kpi_label = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    fill_rec_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_item_1 = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_item_1 = Font(name=font_family, size=10, bold=True, color="166534")

    fill_item_2 = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_item_2 = Font(name=font_family, size=10, bold=True, color="92400E")

    fill_item_3 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    font_item_3 = Font(name=font_family, size=10, bold=True, color="1E40AF")

    status_styles = {
        "정상운행": {"fill": PatternFill(start_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="166534")},
        "유지중": {"fill": PatternFill(start_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="166534")},
        "자사소유": {"fill": PatternFill(start_color="F3E8FF", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="6B21A8")},
        "양수완료": {"fill": PatternFill(start_color="DBEAFE", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="1E40AF")},
        "만기해지": {"fill": PatternFill(start_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="475569")},
        "양도완료": {"fill": PatternFill(start_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="475569")},
        "중도해지": {"fill": PatternFill(start_color="FEE2E2", fill_type="solid"), "font": Font(name=font_family, size=10, bold=True, color="991B1B")},
    }

    thin_border_side = Side(style='thin', color='CBD5E1')
    thick_bottom_side = Side(style='medium', color='1E293B')
    double_bottom_side = Side(style='double', color='1E293B')

    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_bottom_side)

    # SHEET 1: 2026년도_법인차량_총괄자산대장
    ws1 = wb.active
    ws1.title = "2026년도_법인차량_총괄자산대장"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Row 1)
    ws1.merge_cells("A1:S1")
    title_cell = ws1["A1"]
    title_cell.value = "[㈜대상기업] 2026년도 법인차량 총괄자산대장 [외부감사 및 IPO 제출용]"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    # Subtitle (Row 2)
    ws1.cell(row=2, column=1, value="※ 스냅샷 기준일자: 2026년 12월 31일 기준 | 단위: 원 (VAT 포함) | 작성부서: 경영지원본부 총무팀").font = font_subtitle
    ws1.row_dimensions[2].height = 18

    # Summary Section Title (Row 4)
    ws1.merge_cells("A4:S4")
    sum_title_cell = ws1["A4"]
    sum_title_cell.value = "📊 2026년도 법인차량 자산 기말 재무 구분 요약 (Executive Financial Summary)"
    sum_title_cell.font = font_kpi_hdr
    sum_title_cell.fill = fill_kpi_hdr
    sum_title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 20):
        ws1.cell(row=4, column=c).border = border_cell
    ws1.row_dimensions[4].height = 24

    # KPI Summary Cards (Rows 5-6) - Updated calculated values from docx!
    kpis = [
        ("총 관리자산", "10 개 자산", "A", "C"),
        ("정상운행 자산", "8 대 (렌트5/리스3)", "D", "F"),
        ("① 렌탈 / 리스 보증금", 95549000, "G", "I"),
        ("② 월 렌탈 / 리스료", 10684720, "J", "L"),
        ("③ 연간 총 리스/렌트비", 128216640, "M", "O"),
        ("④ 해지 / 승계 완료 자산", "2 대 (양도 완료)", "P", "S"),
    ]

    for label, val, c_start, c_end in kpis:
        ws1.merge_cells(f"{c_start}5:{c_end}5")
        l_cell = ws1[f"{c_start}5"]
        l_cell.value = label
        l_cell.font = font_kpi_label
        l_cell.fill = fill_kpi_label
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        start_col = openpyxl.utils.column_index_from_string(c_start)
        end_col = openpyxl.utils.column_index_from_string(c_end)
        for col_idx in range(start_col, end_col + 1):
            ws1.cell(row=5, column=col_idx).border = border_cell

        ws1.merge_cells(f"{c_start}6:{c_end}6")
        v_cell = ws1[f"{c_start}6"]
        v_cell.value = val
        v_cell.font = font_kpi_val
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(val, (int, float)):
            v_cell.number_format = "#,#0"

        for col_idx in range(start_col, end_col + 1):
            ws1.cell(row=6, column=col_idx).border = border_cell

    ws1.row_dimensions[5].height = 20
    ws1.row_dimensions[6].height = 24

    # Master Table Headers (Row 8)
    headers = [
        "순번", "권역", "자산 구분", "차종 (모델명)", "차량번호",
        "금융사 (렌트/리스)", "주 운행자 / 부서", "최초계약일", "임대기간 (시작-종료)",
        "지급 주기", "매월 납부일", "은 행", "계 좌 번 호", "예 금 주",
        "보 증 금", "임 대 료", "약정 주행거리", "당해년도 상태", "비 고"
    ]

    ws1.row_dimensions[8].height = 28
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws1.cell(row=8, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin_border_side, right=thin_border_side, top=thick_bottom_side, bottom=thick_bottom_side)

    # Master Data Rows (Rows 9-18) - 100% matched to Word Contract Notes (docx)!
    vehicles_docx_matched = [
        (1, "가산", "해지자산", "GV70 (141호8727)", "141호8727", "현대캐피탈", "임직원 업무용", "2021-12-27", "2021/12/27 - 2026/12/26", "월세", "자동이체", "법인 지정계좌", "현대캐피탈", "현대캐피탈㈜", 0, 0, "연 2만~3만 km", "양도완료", "2024.08.02 타사/타인 양도 처리 완료"),
        (2, "강남", "임차자산", "K8 (289수3930)", "289수3930", "현대캐피탈", "임직원 업무용 (전략본부)", "2022-01-04", "2022/01/04 - 2027/01/04", "월세", "자동이체", "법인 지정계좌", "현대캐피탈", "현대캐피탈㈜", 0, 555600, "연 2만~3만 km", "정상운행", "정상 운행 중 (운용리스)"),
        (3, "가산", "임차자산", "그랜저 (141하9479)", "141하9479", "현대캐피탈", "임직원 업무용 (최광일)", "2022-01-04", "2022/01/04 - 2027/01/04", "월세", "자동이체", "법인 지정계좌", "현대캐피탈", "현대캐피탈㈜", 0, 646580, "연 2만~3만 km", "정상운행", "정상 운행 중 (장기렌트)"),
        (4, "광명", "임차자산", "GV80 (103하8547)", "103하8547", "DGB(IM캐피탈)", "임직원 업무용 (경영기획실)", "2022-02-28", "2022/02/28 - 2027/02/28", "월세", "자동이체", "법인 지정계좌", "DGB(IM캐피탈)", "DGB캐피탈㈜", 17360000, 1342660, "연 2만~3만 km", "양수완료", "2025.11.07 승계/양수 완료 (장기렌트)"),
        (5, "강남", "임차자산", "벤츠S클래스 (281가8991)", "281가8991", "하나캐피탈", "임직원 업무용 (이종탁 대표님)", "2022-03-14", "2022/03/14 - 2027/03/11", "월세", "자동이체", "법인 지정계좌", "하나캐피탈", "하나캐피탈㈜", 51459000, 3167300, "연 2만~3만 km", "정상운행", "대표이사 의전 차량 (운용리스)"),
        (6, "대전", "임차자산", "카니발 (269더5669)", "269더5669", "현대캐피탈", "임직원 업무용 (GL실/이준민)", "2022-03-23", "2022/03/23 - 2027/03/23", "월세", "자동이체", "법인 지정계좌", "현대캐피탈", "현대캐피탈㈜", 0, 984000, "연 2만~3만 km", "정상운행", "정상 운행 중 (운용리스)"),
        (7, "판교", "임차자산", "스포티지 (167호2430)", "167호2430", "우리캐피탈", "임직원 업무용 (김철병)", "2022-06-24", "2022/06/24 - 2027/06/23", "월세", "자동이체", "법인 지정계좌", "우리캐피탈", "우리캐피탈㈜", 0, 752500, "연 2만~3만 km", "정상운행", "정상 운행 중 (장기렌트)"),
        (8, "강남", "임차자산", "GV80 (197호3290)", "197호3290", "하나캐피탈", "임직원 업무용 (정명훈)", "2022-09-28", "2022/09/28 - 2027/09/27", "월세", "자동이체", "법인 지정계좌", "하나캐피탈", "하나캐피탈㈜", 26730000, 1449580, "연 2만~3만 km", "정상운행", "정상 운행 중 (장기렌트)"),
        (9, "대전", "해지자산", "GV70 (172하6158)", "172하6158", "농협캐피탈", "임직원 업무용", "2022-11-10", "2022/11/10 - 2027/11/09", "월세", "자동이체", "법인 지정계좌", "농협캐피탈", "농협캐피탈㈜", 0, 0, "연 2만~3만 km", "양도완료", "2024.08.02 타사/타인 양도 처리 완료"),
        (10, "강남", "임차자산", "아우디A8 (120너2842)", "120너2842", "BNK캐피탈", "임직원 업무용 (임은희)", "2024-11-14", "2024/11/14 - 2029/11/13", "월세", "자동이체", "법인 지정계좌", "BNK캐피탈", "BNK캐피탈㈜", 0, 1787000, "연 2만~3만 km", "정상운행", "정상 운행 중 (운용리스)"),
    ]

    start_row = 9
    for idx, v_data in enumerate(vehicles_docx_matched):
        r_num = start_row + idx
        ws1.row_dimensions[r_num].height = 22
        fill_curr = fill_zebra if idx % 2 == 1 else fill_white

        for c_offset, val in enumerate(v_data):
            col_idx = 1 + c_offset
            cell = ws1.cell(row=r_num, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_curr
            cell.border = border_cell

            if c_offset in [0, 1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 13, 16, 17]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_offset in [3, 12, 18]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_offset in [14, 15]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,#0"
                cell.font = font_bold if val > 0 else font_body

            if c_offset == 17:
                st_info = status_styles.get(val)
                if st_info:
                    cell.fill = st_info["fill"]
                    cell.font = st_info["font"]

    # Grand Total Row
    tot_row = start_row + len(vehicles_docx_matched)
    ws1.row_dimensions[tot_row].height = 26
    ws1.merge_cells(f"A{tot_row}:N{tot_row}")
    tot_label = ws1[f"A{tot_row}"]
    tot_label.value = "합  계 (유효 운행차량 기준)"
    tot_label.font = font_bold
    tot_label.alignment = Alignment(horizontal="center", vertical="center")
    
    for c in range(1, 20):
        cell = ws1.cell(row=tot_row, column=c)
        cell.fill = fill_total
        cell.border = border_total
        if c == 15:
            cell.value = f'=SUMIF(R{start_row}:R{tot_row-1}, "정상운행", O{start_row}:O{tot_row-1}) + SUMIF(R{start_row}:R{tot_row-1}, "양수완료", O{start_row}:O{tot_row-1})'
            cell.number_format = "#,#0"
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c == 16:
            cell.value = f'=SUMIF(R{start_row}:R{tot_row-1}, "정상운행", P{start_row}:P{tot_row-1}) + SUMIF(R{start_row}:R{tot_row-1}, "양수완료", P{start_row}:P{tot_row-1})'
            cell.number_format = "#,#0"
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Set EXPLICIT Column Widths (matching Real Estate Col A = 8.0)
    v_col_widths = {
        'A': 8.0,   # 순번 (Exact 8.0)
        'B': 10.0,  # 권역
        'C': 12.0,  # 자산 구분
        'D': 25.0,  # 차종
        'E': 16.0,  # 차량번호
        'F': 18.0,  # 금융사
        'G': 28.0,  # 운행자/부서
        'H': 14.0,  # 최초계약일
        'I': 28.0,  # 계약기간
        'J': 10.0,  # 지급주기
        'K': 12.0,  # 납부일
        'L': 16.0,  # 은행
        'M': 22.0,  # 계좌번호
        'N': 16.0,  # 예금주
        'O': 18.0,  # 보증금
        'P': 16.0,  # 임대료
        'Q': 18.0,  # 약정주행거리
        'R': 14.0,  # 당해년도 상태
        'S': 38.0,  # 비고
    }
    for col_let, w in v_col_widths.items():
        ws1.column_dimensions[col_let].width = w

    # SHEET 2: 2024-2026_차량변동이력
    ws2 = wb.create_sheet(title="2024-2026_차량변동이력")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    t2_cell = ws2["A1"]
    t2_cell.value = " 📜 [㈜대상기업] 법인차량 3개년 변동 이력 타임라인 (2024 ~ 2026)"
    t2_cell.font = font_title
    t2_cell.fill = fill_title
    t2_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 36

    v_hist_headers = ["연도", "변동 일자", "권역", "차종 및 차량번호", "변동 구분", "보증금/선납금 (원)", "월 렌탈/리스료 (원)", "세부 변동 이력 및 사유"]
    ws2.row_dimensions[3].height = 28
    for idx, h in enumerate(v_hist_headers, start=1):
        c = ws2.cell(row=3, column=idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=thin_border_side, right=thin_border_side, top=thick_bottom_side, bottom=thick_bottom_side)

    v_history = [
        ("2024년", "2024.08.02", "가산", "GV70 (141호8727)", "양도 완료", 0, -1096150, "GV70 141호8727 타사/타인 양도 처리 완료"),
        ("2024년", "2024.08.02", "대전", "GV70 (172하6158)", "양도 완료", -17445000, -1064000, "GV70 172하6158 타사/타인 양도 및 보증금 1,744.5만원 환수 완료"),
        ("2024년", "2024.11.14", "강남", "아우디A8 (120너2842)", "신규 리스", 0, 1787000, "아우디A8 운용리스 신규 계약 (임직원 업무용)"),
        ("2025년", "2025.11.07", "광명", "GV80 (103하8547)", "승계 양수", 17360000, 1342660, "GV80 103하8547 승계/양수 완료 (경영기획실 업무용)"),
    ]

    last_timeline_row = 3 + len(v_history)
    for idx, h_data in enumerate(v_history):
        r_num = 4 + idx
        ws2.row_dimensions[r_num].height = 22
        fill_curr = fill_zebra if idx % 2 == 1 else fill_white

        for c_off, val in enumerate(h_data):
            col = 1 + c_off
            cell = ws2.cell(row=r_num, column=col, value=val)
            cell.font = font_body
            cell.fill = fill_curr
            cell.border = border_cell

            if c_off in [0, 1, 2, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_off == 3 or c_off == 7:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_off in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,#0"

    rec_start_row = last_timeline_row + 2
    ws2.merge_cells(f"A{rec_start_row}:H{rec_start_row}")
    r_hdr_cell = ws2[f"A{rec_start_row}"]
    r_hdr_cell.value = "📊 회계/재무 검증용 차량 보증금/리스료 3개년 현금흐름 및 기말 잔액 대사표 (Reconciliation Table)"
    r_hdr_cell.font = font_kpi_hdr
    r_hdr_cell.fill = fill_rec_hdr
    r_hdr_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 9):
        ws2.cell(row=rec_start_row, column=c).border = border_cell
    ws2.row_dimensions[rec_start_row].height = 26

    rec_col_hdr_row = rec_start_row + 1
    ws2.merge_cells(f"A{rec_col_hdr_row}:D{rec_col_hdr_row}")
    ws2[f"A{rec_col_hdr_row}"].value = "회계 대사 구분 항목"
    ws2[f"A{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells(f"E{rec_col_hdr_row}:F{rec_col_hdr_row}")
    ws2[f"E{rec_col_hdr_row}"].value = "금액 (원)"
    ws2[f"E{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells(f"G{rec_col_hdr_row}:H{rec_col_hdr_row}")
    ws2[f"G{rec_col_hdr_row}"].value = "회계장부(BS) 및 현금흐름 대사 설명"
    ws2[f"G{rec_col_hdr_row}"].alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 9):
        cell = ws2.cell(row=rec_col_hdr_row, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_cell
    ws2.row_dimensions[rec_col_hdr_row].height = 24

    reconciliations = [
        ("① 기간 중 현금 환수/회수 보증금 총액", 17445000, "2024년 GV70(172하6158) 타사 양도에 따른 보증금 1,744.5만 원 회수액"),
        ("   - 2024년 중 보증금 환수/회수액", 17445000, "GV70 172하6158 타사 양도 완료 입금액"),
        ("② 기간 중 순 보증금 현금 변동액 (Net Cash Flow)", -85000, "2025년 GV80(103하) 승계 집행 1,736만 원 - 총 회수액 1,744.5만 원 = 순입금 8.5만 원"),
        ("③ 2026.12.31 기말 재무상태표상 차량보증금 장부 잔액", 95549000, "재무상태표(BS) 차량 임차/리스보증금 장부 잔액과 1:1 일치"),
        ("   - 정상운행 및 승계 8대 보증금 장부 잔액 소계", 95549000, "GV80(17.36백), 벤츠S(51.46백), GV80(26.73백) 3대 보증금 합계"),
    ]

    r_item_base_row = rec_col_hdr_row + 1
    for idx, r_data in enumerate(reconciliations):
        r_num = r_item_base_row + idx
        ws2.row_dimensions[r_num].height = 22
        item_title, item_amt, item_desc = r_data

        item_fill = fill_white
        item_font = font_body
        if idx == 0:
            item_fill, item_font = fill_item_1, font_item_1
        elif idx == 2:
            item_fill, item_font = fill_item_2, font_item_2
        elif idx == 3:
            item_fill, item_font = fill_item_3, font_item_3

        ws2.merge_cells(f"A{r_num}:D{r_num}")
        c_item = ws2[f"A{r_num}"]
        c_item.value = item_title
        c_item.font = item_font
        c_item.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws2.merge_cells(f"E{r_num}:F{r_num}")
        c_amt = ws2[f"E{r_num}"]
        c_amt.value = item_amt
        c_amt.font = item_font
        c_amt.alignment = Alignment(horizontal="right", vertical="center")
        c_amt.number_format = "#,#0"

        ws2.merge_cells(f"G{r_num}:H{r_num}")
        c_desc = ws2[f"G{r_num}"]
        c_desc.value = item_desc
        c_desc.font = item_font if item_fill != fill_white else font_body
        c_desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for col_idx in range(1, 9):
            ws2.cell(row=r_num, column=col_idx).fill = item_fill
            ws2.cell(row=r_num, column=col_idx).border = border_cell

    s2_col_widths = {
        'A': 10.0, 'B': 14.0, 'C': 10.0, 'D': 35.0,
        'E': 14.0, 'F': 20.0, 'G': 24.0, 'H': 68.0
    }
    for col_let, w in s2_col_widths.items():
        ws2.column_dimensions[col_let].width = w

    filename = "[외감_IPO대비]_2026년도_법인차량_총괄자산대장_[㈜대상기업].xlsx"
    local_path = os.path.join(r"C:\Users\User\OneDrive\문서\GitHub\FoxConnect-AX\output", filename)
    gdrive_path = os.path.join(r"G:\내 드라이브\[FoxConnect]\[총무]업무\02_차량_자산관리\00_연도별_차량_총괄자산대장", filename)

    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    wb.save(local_path)
    print(f"✅ Docx Matched Vehicle Master Excel saved to Local: {local_path}")

    if os.path.exists(os.path.dirname(gdrive_path)):
        try:
            wb.save(gdrive_path)
            print(f"✅ Docx Matched Vehicle Master Excel saved to GDrive: {gdrive_path}")
        except Exception as e:
            print("GDrive save skipped:", e)

if __name__ == "__main__":
    build_docx_matched_vehicle_excel()
