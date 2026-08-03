"""
Universal Multi-Tab Real Estate Master Excel Engine (Audit & IPO Compliant)
Refactored to 3-Tier Financial Breakdown & Accountant-Friendly Reconciliation Format
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import FONT_FAMILY

def generate_real_estate_excel(base_dir, company_name="[㈜대상기업]", snapshot_date="2026년 12월 31일"):
    """
    Generates an Audit & IPO compliant multi-tab Real Estate Master Excel Register (.xlsx).
    """
    target_dir = os.path.join(base_dir, r"01_부동산_자산관리\00_연도별_부동산_총괄자산대장")
    os.makedirs(target_dir, exist_ok=True)
    
    file_name = f"[외감_IPO대비]_2026년도_부동산_총괄자산대장_[㈜대상기업].xlsx"
    excel_path = os.path.join(target_dir, file_name)
    local_path = os.path.join(base_dir, "output", file_name)
    os.makedirs(os.path.join(base_dir, "output"), exist_ok=True)
    
    print(f"[RealEstateEngine] Generating Master Real Estate Register at:\n  {excel_path}\n  {local_path}")
    
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: 2026년도_부동산_총괄자산대장 (3단계 구분 재무 요약 규격)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "2026년도_부동산_총괄자산대장"
    ws1.views.sheetView[0].showGridLines = True

    font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name=FONT_FAMILY, size=10, italic=True, color="475569")
    font_header = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    font_sub_header = Font(name=FONT_FAMILY, size=11, bold=True, color="1E293B")
    font_body = Font(name=FONT_FAMILY, size=10, color="0F172A")
    font_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="0F172A")

    fill_title = PatternFill(start_color="1E293B", fill_type="solid")
    fill_header = PatternFill(start_color="334155", fill_type="solid")
    fill_summary_hdr = PatternFill(start_color="E2E8F0", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", fill_type="solid")

    status_styles = {
        "정상유지": {"fill": PatternFill(start_color="DCFCE7", fill_type="solid"), "font": Font(name=FONT_FAMILY, size=10, bold=True, color="166534")},
        "소유권보유": {"fill": PatternFill(start_color="F3E8FF", fill_type="solid"), "font": Font(name=FONT_FAMILY, size=10, bold=True, color="6B21A8")},
        "만기해지": {"fill": PatternFill(start_color="F1F5F9", fill_type="solid"), "font": Font(name=FONT_FAMILY, size=10, bold=True, color="475569")},
        "중도해지": {"fill": PatternFill(start_color="FEE2E2", fill_type="solid"), "font": Font(name=FONT_FAMILY, size=10, bold=True, color="991B1B")},
        "재계약검토": {"fill": PatternFill(start_color="FEF3C7", fill_type="solid"), "font": Font(name=FONT_FAMILY, size=10, bold=True, color="92400E")}
    }

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_thick = Side(border_style="medium", color="1E293B")
    border_double = Side(border_style="double", color="1E293B")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thick)
    total_border = Border(top=border_thin, bottom=border_double)

    # Title Block
    ws1.merge_cells("A1:P1")
    title_cell = ws1["A1"]
    title_cell.value = f" {company_name} 2026년도 부동산 총괄자산대장 [외부감사 및 IPO 제출용]"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 40

    ws1["A2"].value = f"※ 스냅샷 기준일자: {snapshot_date} 기준 | 단위: 원 (VAT 포함) | 작성부서: 경영지원본부 총무팀"
    ws1["A2"].font = font_subtitle
    ws1.row_dimensions[2].height = 20

    # Executive Summary Box (Rows 4~6)
    ws1.merge_cells("A4:P4")
    ws1["A4"].value = "📊 2026년도 부동산 자산 기말 재무 구분 요약 (Executive Financial Summary)"
    ws1["A4"].font = font_sub_header
    ws1["A4"].fill = fill_summary_hdr
    ws1["A4"].alignment = Alignment(horizontal="left", vertical="center")

    summary_headers = ["총 관리자산", "정상유지 자산", "① 순수 업무용 임차보증금", "② 주거용 전세보증금", "③ 소유자산 취득가액", "④ 연간 총 임차 고정비용", "해지 완료 자산"]
    summary_values = ["17 개 자산", "9 개 사업장", 327000000, 1200000000, 1022660000, 907007788, "6 개 자산"]

    ws1.merge_cells("A5:B5"); ws1["A5"].value = summary_headers[0]
    ws1.merge_cells("C5:D5"); ws1["C5"].value = summary_headers[1]
    ws1.merge_cells("E5:F5"); ws1["E5"].value = summary_headers[2]
    ws1.merge_cells("G5:H5"); ws1["G5"].value = summary_headers[3]
    ws1.merge_cells("I5:J5"); ws1["I5"].value = summary_headers[4]
    ws1.merge_cells("K5:M5"); ws1["K5"].value = summary_headers[5]
    ws1.merge_cells("N5:P5"); ws1["N5"].value = summary_headers[6]

    ws1.merge_cells("A6:B6"); ws1["A6"].value = summary_values[0]
    ws1.merge_cells("C6:D6"); ws1["C6"].value = summary_values[1]
    ws1.merge_cells("E6:F6"); ws1["E6"].value = summary_values[2]
    ws1.merge_cells("G6:H6"); ws1["G6"].value = summary_values[3]
    ws1.merge_cells("I6:J6"); ws1["I6"].value = summary_values[4]
    ws1.merge_cells("K6:M6"); ws1["K6"].value = summary_values[5]
    ws1.merge_cells("N6:P6"); ws1["N6"].value = summary_values[6]

    for r in [5, 6]:
        for c in range(1, 17):
            cell = ws1.cell(r, c)
            cell.font = font_bold if r == 6 else font_body
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = box_border
            if r == 5:
                cell.fill = PatternFill(start_color="F1F5F9", fill_type="solid")
            if r == 6 and c in [5, 6, 7, 8, 9, 10, 11, 12, 13]:
                cell.number_format = "#,##0"

    ws1.row_dimensions[4].height = 24
    ws1.row_dimensions[5].height = 22
    ws1.row_dimensions[6].height = 26

    # Main Table Headers (16 Columns)
    headers = [
        "순번", "권역", "자산 구분", "물건지 및 호수명", "상세 주소", 
        "임대인 (전대인)", "임차인 (전차인)", "최초계약일", "임대기간 (시작-종료)", "지급 주기",
        "보증금 (원)", "월 임대료 (원)", "연간 환산 비용 (원)", "계약면적 (㎡)", "전용면적 (㎡)",
        "2026.12.31 기준 상태"
    ]

    ws1.row_dimensions[8].height = 28
    for col_num, h_text in enumerate(headers, 1):
        cell = ws1.cell(8, col_num)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    raw_data = [
        (1, "가산", "임차자산", "대륭포스트타워6차 402~403호", "서울시 금천구 벚꽃로 298 402~403호", "㈜엠씨에스솔루션", "[㈜대상기업]", "2024-02-29", "2026/02/28 - 2027/02/28", "월세", 46000000, 5060000, 60720000, 218.7, 218.7, "정상유지"),
        (2, "가산", "임차자산", "대륭포스트타워6차 1510호", "서울시 금천구 벚꽃로 298 1510호", "㈜라미나알앤디", "[㈜대상기업]", "2026-04-29", "2026/04/29 - 2028/04/28", "월세", 21000000, 2100000, 25200000, 105.8, 105.8, "정상유지"),
        (3, "광명", "소유자산", "GIDC 지식산업센터 1212호", "경기도 광명시 일직로 43 A동 1212호", "하늘기획(매도인)", "[㈜대상기업]", "2022-01-13", "2022/01/13 - 소유권보유", "소유", 494995000, 0, 0, 189.45, 94.76, "소유권보유"),
        (4, "광명", "소유자산", "GIDC 지식산업센터 1213호", "경기도 광명시 일직로 43 A동 1213호", "하늘기획(매도인)", "[㈜대상기업]", "2022-01-13", "2022/01/13 - 소유권보유", "소유", 527665000, 0, 0, 197.04, 98.56, "소유권보유"),
        (5, "대전", "임차자산", "도룡동 하우스디어반 B721호", "대전시 유성구 도룡동 4-9 B동 721호", "정근호", "[㈜대상기업]", "2022-04-07", "2026/04/08 - 2027/04/07", "월세", 10000000, 790000, 9480000, 35.76, 35.76, "정상유지"),
        (6, "대전", "임차자산", "리저브 도룡동 385-28", "대전시 유성구 도룡동 385-28", "김순미", "[㈜대상기업]", "2023-02-28", "2023/02/28 - 2028/02/28", "월세", 100000000, 6050000, 72600000, 327.27, 327.27, "정상유지"),
        (7, "대전", "임차자산", "대전센터 골프존 204호", "대전시 유성구 엑스포로97번길 40 204호", "(주)골프존홀딩스", "[㈜대상기업]", "2021-08-17", "2026/03/26 - 2027/05/31", "월세", 50000000, 8874789, 106497468, 493.22, 493.22, "정상유지"),
        (8, "대전", "임차자산", "대전센터 골프존 2층 상담실", "대전시 유성구 엑스포로97번길 40 2층 상담실", "(주)골프존홀딩스", "[㈜대상기업]", "2021-08-17", "2026/03/26 - 2027/05/31", "월세", 0, 475860, 5710320, 26.44, 26.44, "정상유지"),
        (9, "판교", "임차자산", "이레빌딩 3층", "경기도 성남시 분당구 운중동 1017-2 3층", "화코스텍인터내셔널", "[㈜대상기업]", "2023-01-31", "2023/02/28 - 2028/02/28", "월세", 100000000, 11000000, 132000000, 396.69, 396.69, "정상유지"),
        (10, "대전", "임차자산", "스마트시티 2501호", "대전시 유성구 도룡동 4-6 2501호", "김동하", "[㈜대상기업]", "2022-02-16", "2022/02/16 - 2024/11/03", "전세", 1200000000, 0, 0, 134.94, 134.94, "정상유지"),
        (11, "대전", "임차자산", "갑동 K-데이터센터", "대전시 유성구 갑동 388-1", "㈜에이아이데이타", "[㈜대상기업]", "2024-12-23", "2024/12/23 - 2026/12/22", "분기납", 0, 0, 520000000, 0, 0, "재계약검토"),
        (12, "대전", "해지자산", "스마트시티상가 115호", "대전시 유성구 도룡동 4-6 205동 115호", "강남규", "[㈜대상기업]", "2023-11-10", "2025/11/10 - 2026/12/10", "월세", 30000000, 1100000, 0, 36.55, 36.55, "만기해지"),
        (13, "대전", "해지자산", "스마트시티상가 113호", "대전시 유성구 도룡동 4-6 205동 113호", "강남규", "[㈜대상기업]", "2023-11-10", "2025/11/10 - 2026/12/10", "월세", 80000000, 1760000, 0, 79.33, 79.33, "만기해지"),
        (14, "강남", "해지자산", "강남본점 (도곡로1길 23 전층)", "서울시 강남구 도곡로1길 23(역삼동)", "유한회사 청송", "[㈜대상기업]", "2024-11-01", "2024/11/01 - 2026/03/31", "월세", 200000000, 16170000, 0, 822.8, 822.8, "만기해지"),
        (15, "대전", "해지자산", "KCC웰츠텔 101동 1202호", "대전시 유성구 도룡동 4-30 1202호", "이선영", "[㈜대상기업]", "2022-04-11", "2025/04/14 - 2026/05/19", "월세", 20000000, 650000, 0, 26.11, 26.11, "만기해지"),
        (16, "대전", "해지자산", "스마트시티상가 209호", "대전시 유성구 도룡동 4-6 209호", "김지영", "[㈜대상기업]", "2024-01-31", "2026/02/28 - 2026/06/30", "월세", 10000000, 660000, 0, 38.01, 38.01, "만기해지"),
        (17, "광명", "해지자산", "GIDC 지식산업센터 1214~1215호", "경기도 광명시 일직로 43 A동 1214~1215호", "하진우", "[㈜대상기업]", "2022-06-22", "2025/08/01 - 2026/07/31", "월세", 25000000, 2750000, 0, 139.24, 139.24, "만기해지")
    ]

    start_row = 9
    for idx, row_item in enumerate(raw_data, start=start_row):
        ws1.row_dimensions[idx].height = 22
        fill_curr = fill_zebra if idx % 2 == 1 else fill_white
        
        for col_idx, val in enumerate(row_item, 1):
            cell = ws1.cell(idx, col_idx)
            cell.value = val
            cell.font = font_body
            cell.fill = fill_curr
            cell.border = box_border

            if col_idx in [1, 2, 3, 8, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [4, 5, 6, 7, 9]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [11, 12, 13, 14, 15]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)) and col_idx in [11, 12, 13]:
                    cell.number_format = "#,##0"
                elif isinstance(val, (int, float)) and col_idx in [14, 15]:
                    cell.number_format = "#,##0.0"

            if col_idx == 16:
                st_name = str(val)
                if st_name in status_styles:
                    cell.fill = status_styles[st_name]["fill"]
                    cell.font = status_styles[st_name]["font"]
                cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = start_row + len(raw_data)
    ws1.row_dimensions[total_row].height = 26
    
    ws1.cell(total_row, 1).value = "합  계 (유효 임차자산 기준)"
    ws1.cell(total_row, 1).font = font_bold
    ws1.cell(total_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
    
    cell_dep = ws1.cell(total_row, 11)
    cell_dep.value = f"=SUMIF(P9:P{total_row-1}, \"정상유지\", K9:K{total_row-1})"
    cell_dep.font = font_bold
    cell_dep.number_format = "#,##0"
    cell_dep.alignment = Alignment(horizontal="right", vertical="center")
    
    cell_rent = ws1.cell(total_row, 12)
    cell_rent.value = f"=SUMIF(P9:P{total_row-1}, \"정상유지\", L9:L{total_row-1})"
    cell_rent.font = font_bold
    cell_rent.number_format = "#,##0"
    cell_rent.alignment = Alignment(horizontal="right", vertical="center")

    cell_annual = ws1.cell(total_row, 13)
    cell_annual.value = f"=SUM(M9:M{total_row-1})"
    cell_annual.font = font_bold
    cell_annual.number_format = "#,##0"
    cell_annual.alignment = Alignment(horizontal="right", vertical="center")

    for col in range(1, 17):
        cell = ws1.cell(total_row, col)
        cell.border = total_border
        cell.fill = PatternFill(start_color="F1F5F9", fill_type="solid")

    # SHEET 2: 2024-2026_자산변동이력 (Accountant-Friendly Vertical Layout)
    ws2 = wb.create_sheet(title="2024-2026_자산변동이력")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value = f" 📜 {company_name} 부동산 자산 3개년 변동 이력 타임라인 (2024 ~ 2026)"
    t2.font = font_title
    t2.fill = fill_title
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 36

    h2 = ["연도", "변동 일자", "권역", "물건지 및 호수명", "변동 구분", "보증금 (원)", "월 임대료 / 지출 (원)", "세부 변동 이력 및 사유"]
    ws2.row_dimensions[3].height = 26
    for c_idx, h_t in enumerate(h2, 1):
        c = ws2.cell(3, c_idx)
        c.value = h_t
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_border

    history_data = [
        ("2024년", "2024.01.31", "대전", "스마트시티상가 209호", "최초 임차", 10000000, 660000, "대전 원격학원 용도 최초 계약 체결"),
        ("2024년", "2024.02.29", "가산", "대륭포스트타워6차 402~403호", "최초 임차", 46000000, 5060000, "문항개발실 거점 사무실 최초 임차 (임대인: ㈜엠씨에스솔루션)"),
        ("2024년", "2024.11.01", "강남", "강남본점 (도곡로1길 23 전층)", "최초 임차", 200000000, 16170000, "강남 사옥 통합 거점 최초 임차 (전용 248.9평)"),
        ("2024년", "2024.12.23", "대전", "갑동 388-1 K-데이터센터", "최초 임차", 0, "분기 130,000,000", "K-데이터센터 임차 개시 (1/4/7/10월 분기별 1.3억 납부)"),
        ("2025년", "2025.04.17", "대전", "스마트시티 504-604호 (원어민숙소)", "계약 종료", -10000000, -900000, "원어민 숙소 계약 만료 퇴거 및 보증금 1,000만원 반환 완료"),
        ("2025년", "2025.06.01", "대전", "대전센터 골프존 204호/상담실", "연장 계약", 50000000, 9350649, "러닝센터 및 ID센터 임대차 기간 연장 계약"),
        ("2025년", "2025.09.01", "강남", "강남본점 (도곡로1길 23 전층)", "임대인 변경", 200000000, 16170000, "임대인 명의 변경 (박재윤 -> 유한회사 청송)"),
        ("2025년", "2025.11.10", "대전", "딥러닝 스마트시티상가 113/115호", "연장 계약", 110000000, 2860000, "딥러닝센터 2년 임대차 연장 계약 체결"),
        ("2026년", "2026.02.28", "가산", "대륭포스트타워6차 402~403호", "묵시적 갱신", 46000000, 5060000, "관리자 확인: 2026.02.28 이후 묵시적 자동연장 계속 유지 중"),
        ("2026년", "2026.03.26", "대전", "대전센터 골프존 204호/상담실", "연장 계약", 50000000, 9350649, "연장계약서(260326.pdf) 반영 (2027.05.31까지 연장)"),
        ("2026년", "2026.03.31", "강남", "강남본점 (도곡로1길 23 전층)", "만기 해지", -200000000, -16170000, "사옥 거점 이전 완료에 따른 계약 만료 종료 및 보증금 2억원 반환 완료"),
        ("2026년", "2026.04.08", "대전", "도룡동 하우스디어반 B721호", "묵시적 갱신", 10000000, 790000, "관리자 확인: 2026.04.08 이후 묵시적 자동연장 계속 유지 중"),
        ("2026년", "2026.04.29", "가산", "대륭포스트타워6차 1510호", "신규 계약", 21000000, 2100000, "경영지원본부 사무실 신규 임차 계약 체결"),
        ("2026년", "2026.05.19", "대전", "KCC웰츠텔 101동 1202호", "만기 해지", -20000000, -650000, "계약 종료 및 관리비 정산, 보증금 2,000만원 반환 완료"),
        ("2026년", "2026.06.30", "대전", "스마트시티상가 209호", "만기 해지", -10000000, -660000, "원격학원 계약 종료 및 관리비 정산, 보증금 1,000만원 반환 완료"),
        ("2026년", "2026.07.31", "광명", "GIDC 지식산업센터 1214~1215호", "만기 해지", -25000000, -2750000, "임대차 계약 만료에 따른 보증금 2,500만원 반환 완료"),
        ("2026년", "2026.12.10", "대전", "딥러닝 스마트시티상가 113/115호", "만기 해지", -110000000, -2860000, "관리자 지침 반영: 묵시적 연장 없이 2026.12.10 자 만기해지 종료 및 1.1억 반환")
    ]

    for idx, r_item in enumerate(history_data, start=4):
        ws2.row_dimensions[idx].height = 22
        fill_c = fill_zebra if idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(r_item, 1):
            c = ws2.cell(idx, c_idx)
            c.value = val
            c.font = font_body
            c.fill = fill_c
            c.border = box_border
            
            if c_idx == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [2, 3, 5]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [4, 8]:
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in [6, 7]:
                c.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"

    # ACCOUNTANT-FRIENDLY RECONCILIATION SUMMARY TABLE BELOW
    start_sum_row = 22
    
    ws2.merge_cells(f"A{start_sum_row}:H{start_sum_row}")
    hdr_cell = ws2[f"A{start_sum_row}"]
    hdr_cell.value = "📊 회계/재무 검증용 임차보증금 3개년 현금흐름 및 기말 잔액 대사표 (Reconciliation Table)"
    hdr_cell.font = font_sub_hdr
    hdr_cell.fill = fill_summary_hdr
    hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
    hdr_cell.border = header_border
    ws2.row_dimensions[start_sum_row].height = 28

    ws2.merge_cells("A23:D23"); ws2["A23"].value = "회계 대사 구분 항목"
    ws2.merge_cells("E23:F23"); ws2["E23"].value = "금액 (원)"
    ws2.merge_cells("G23:H23"); ws2["G23"].value = "회계장부(BS) 및 현금흐름 대사 설명"
    
    ws2.row_dimensions[23].height = 24
    for c_col in range(1, 9):
        cell = ws2.cell(23, c_col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    summary_rows = [
        ("① 기간 중 현금 회수/환수 보증금 총액", 375000000, "2025~2026년 계약종료에 따른 보증금 통장 입금 회수액 (총 6건)"),
        ("   - 2025년 중 보증금 환수액", 10000000, "스마트시티 504-604호 원어민숙소 환수"),
        ("   - 2026년 중 보증금 환수액", 365000000, "강남본점(2억), KCC(2천), 상가209(1천), GIDC(2.5천), 상가113/115(1.1억)"),
        ("② 기간 중 순 보증금 현금 변동액 (Net Cash Flow)", 62000000, "3개년 신규집행 4.37억 - 총 회수액 3.75억 = 순유출 6,200만 원"),
        ("③ 2026.12.31 기말 재무상태표상 보증금 장부 잔액", 1527000000, "재무상태표(BS) 임차/전세보증금 장부 잔액과 1:1 일치"),
        ("   - 업무용 임차보증금 잔액", 327000000, "가산, 대전, 판교 사업장 5개 소계"),
        ("   - 주거용 전세보증금 잔액", 1200000000, "대전 스마트시티 2501호 전세 보증금")
    ]

    for s_idx, (k_item, v_amt, note_str) in enumerate(summary_rows, start=24):
        ws2.row_dimensions[s_idx].height = 22
        
        ws2.merge_cells(f"A{s_idx}:D{s_idx}")
        ws2.merge_cells(f"E{s_idx}:F{s_idx}")
        ws2.merge_cells(f"G{s_idx}:H{s_idx}")

        c_k = ws2.cell(s_idx, 1)
        c_v = ws2.cell(s_idx, 5)
        c_n = ws2.cell(s_idx, 7)

        c_k.value = k_item
        c_v.value = v_amt
        c_n.value = note_str

        for c_col in range(1, 9):
            cell = ws2.cell(s_idx, c_col)
            cell.border = box_border
            cell.font = font_bold if s_idx in [24, 27, 28] else font_body
            if s_idx == 24:
                cell.fill = PatternFill(start_color="DCFCE7", fill_type="solid")
            elif s_idx == 27:
                cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid")
            elif s_idx == 28:
                cell.fill = PatternFill(start_color="DBEAFE", fill_type="solid")

        c_k.alignment = Alignment(horizontal="left", vertical="center")
        c_n.alignment = Alignment(horizontal="left", vertical="center")
        c_v.alignment = Alignment(horizontal="right" if isinstance(v_amt, (int, float)) else "center", vertical="center")
        if isinstance(v_amt, (int, float)):
            c_v.number_format = "#,##0"

    col_widths = {
        'A': 10,
        'B': 14,
        'C': 10,
        'D': 35,
        'E': 14,
        'F': 20,
        'G': 24,
        'H': 68
    }
    for col_letter, w_val in col_widths.items():
        ws2.column_dimensions[col_letter].width = w_val

    for p in [excel_path, local_path]:
        wb.save(p)
        print(f"✅ RealEstateEngine successfully updated Master Register: {p}")

if __name__ == "__main__":
    from src.config import DEFAULT_BASE_DIR, DEFAULT_COMPANY_NAME
    generate_real_estate_excel(DEFAULT_BASE_DIR, DEFAULT_COMPANY_NAME)
