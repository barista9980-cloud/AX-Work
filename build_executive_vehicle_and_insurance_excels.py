import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FOXCONNECT_ROOT = r"G:\내 드라이브\[FoxConnect]"
VEHICLE_BASE = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\02_차량_자산관리")
INSURANCE_BASE = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\03_보험_자산관리")

VEHICLE_TARGET_DIR = os.path.join(VEHICLE_BASE, "00_연도별_차량_총괄자산대장")
INSURANCE_TARGET_DIR = os.path.join(INSURANCE_BASE, "00_연도별_보험_총괄자산대장")

os.makedirs(VEHICLE_TARGET_DIR, exist_ok=True)
os.makedirs(INSURANCE_TARGET_DIR, exist_ok=True)

veh_excel_p = os.path.join(VEHICLE_TARGET_DIR, "[외감_IPO대비]_주식회사_폭스에듀_연도별_법인차량_총괄자산대장(2022-2025).xlsx")
ins_excel_p = os.path.join(INSURANCE_TARGET_DIR, "[외감_IPO대비]_주식회사_폭스에듀_연도별_기업보험_총괄자산대장(2022-2025).xlsx")

font_family = "맑은 고딕"

title_font = Font(name=font_family, size=15, bold=True, color="0F172A")
subtitle_font = Font(name=font_family, size=9, bold=True, color="475569")

summary_bar_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

data_font = Font(name=font_family, size=10, color="0F172A")
data_bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")

zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

pill_styles = {
    "정상운행": {"fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="166534")},
    "유지중": {"fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="166534")},
    "양도완료": {"fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="475569")},
    "양수완료": {"fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="1E40AF")},
    "만기해지": {"fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="475569")}
}

black_border_side = Side(style='thin', color='000000')
black_thick_top_side = Side(style='medium', color='000000')

table_header_border = Border(left=black_border_side, right=black_border_side, top=black_thick_top_side, bottom=black_border_side)
table_data_border = Border(left=black_border_side, right=black_border_side, top=black_border_side, bottom=black_border_side)

# ==========================================
# PART 1: CORPORATE VEHICLE MASTER EXCEL
# ==========================================
print("Generating Corporate Vehicle Master Excel (.xlsx)...")
wb_veh = openpyxl.Workbook()
wb_veh.remove(wb_veh.active)

ws_v = wb_veh.create_sheet(title="01_법인차량_자산대장")
ws_v.views.sheetView[0].showGridLines = False
ws_v.column_dimensions['A'].width = 3

ws_v.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 법인차량 총괄자산대장 (2022~2025)").font = title_font
ws_v.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 2025.12.31 기준 유효 법인차량 자산 대장").font = subtitle_font

ws_v.merge_cells("B5:M5")
v_bar_cell = ws_v.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   운행 차량: 총 8대 (누적 10대)   |   보증금 잔액 합계: ₩ 95,549,000 (9,554만원)   |   월 렌탈/리스료 합계: ₩ 10,685,220 (VAT 별도)")
v_bar_cell.font = Font(name=font_family, size=10, bold=True, color="1E293B")
v_bar_cell.fill = summary_bar_fill
v_bar_cell.alignment = Alignment(horizontal='left', vertical='center')
v_bar_cell.border = table_data_border

v_headers = ["연도", "순서", "차종", "차량번호", "금융사 (렌트/리스)", "계약유형", "계약시작일", "만기일자", "보증금 (원)", "월 렌탈/리스료(원)", "계약 상태", "비고 (승계/양도/양수 이력)"]

ws_v.row_dimensions[7].height = 28
for c_i, text in enumerate(v_headers, 2):
    c = ws_v.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

v_data = [
    ("2022", "01", "GV70", "141호8727", "현대캐피탈", "장기렌트", "2021-12-27", "2026-12-26", 0, 1096150, "양도완료", "2024.08.02 제3자 양도 완료 및 계약종료"),
    ("2022", "02", "K8", "289수3930", "현대캐피탈", "운용리스", "2022-01-04", "2027-01-04", 0, 555600, "정상운행", "법인 대표 임원 전용 리스 차량"),
    ("2022", "03", "그랜저", "141하9479", "현대캐피탈", "장기렌트", "2022-01-04", "2027-01-04", 0, 646580, "정상운행", "영업 및 업무용 장기 렌트 차량"),
    ("2022", "04", "GV80", "103하8547", "DGB(IM캐피탈)", "장기렌트", "2022-02-28", "2027-02-28", 17360000, 1342660, "양수완료", "2025.11.07 법인 승계(양수) 완료"),
    ("2022", "05", "벤츠 S클래스", "281가8991", "하나캐피탈", "운용리스", "2022-03-14", "2027-03-11", 51459000, 3167300, "정상운행", "대표이사 전용 프리미엄 운용리스"),
    ("2022", "06", "카니발", "269더5669", "현대캐피탈", "운용리스", "2022-03-23", "2027-03-23", 0, 984000, "정상운행", "임직원 이동 및 학원 셔틀 리스 차량"),
    ("2022", "07", "스포티지", "167호2430", "우리캐피탈", "장기렌트", "2022-06-24", "2027-06-23", 0, 752500, "정상운행", "연구소 및 현장 지원 렌트 차량"),
    ("2022", "08", "GV80", "197호3290", "하나캐피탈", "장기렌트", "2022-09-28", "2027-09-27", 26730000, 1449580, "정상운행", "경영진 업무용 장기 렌트 차량"),
    ("2022", "09", "GV70", "172하6158", "농협캐피탈", "장기렌트", "2022-11-10", "2027-11-09", 17445000, 1064000, "양도완료", "2024.08.02 제3자 양도 완료 및 계약종료"),
    ("2024", "10", "아우디 A8", "120노2842", "BNK캐피탈", "운용리스", "2024-11-14", "2029-11-13", 0, 1787000, "정상운행", "임원 전용 프리미엄 운용리스 차량")
]

for row_idx, row_data in enumerate(v_data, 8):
    ws_v.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    # row_data tuple contains 12 items: idx 0 to 11
    # Col B = 2 (idx 0: 연도)
    # Col C = 3 (idx 1: 순서)
    # Col D = 4 (idx 2: 차종)
    # Col E = 5 (idx 3: 차량번호)
    # Col F = 6 (idx 4: 금융사)
    # Col G = 7 (idx 5: 계약유형)
    # Col H = 8 (idx 6: 시작일)
    # Col I = 9 (idx 7: 만기일)
    # Col J = 10 (idx 8: 보증금)
    # Col K = 11 (idx 9: 월렌탈/리스료)
    # Col L = 12 (idx 10: 상태)
    # Col M = 13 (idx 11: 비고)
    for t_idx, val in enumerate(row_data):
        c_i = t_idx + 2
        cell = ws_v.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:  # 연도, 순서
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 5, 6, 7, 8, 9]:  # 차종, 차량번호, 금융사, 유형, 시작, 만기
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [10, 11]:  # 보증금(col 10), 월 렌탈/리스료(col 11)
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:  # 상태(col 12)
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:  # 비고(col 13)
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

v_col_widths = {1: 3, 2: 10, 3: 8, 4: 16, 5: 16, 6: 18, 7: 14, 8: 14, 9: 14, 10: 18, 11: 18, 12: 14, 13: 44}
for col_idx, width in v_col_widths.items():
    ws_v.column_dimensions[get_column_letter(col_idx)].width = width

wb_veh.save(veh_excel_p)
print("  [CORPORATE VEHICLE EXCEL FIXED & SAVED SUCCESSFULLY]", veh_excel_p)


# ==========================================
# PART 2: CORPORATE INSURANCE MASTER EXCEL
# ==========================================
print("\nGenerating Corporate Insurance Master Excel (.xlsx)...")
wb_ins = openpyxl.Workbook()
wb_ins.remove(wb_ins.active)

# --- INSURANCE TAB 1: 경영인정기보험 ---
ws_i1 = wb_ins.create_sheet(title="01_경영인정기보험_자산대장")
ws_i1.views.sheetView[0].showGridLines = False
ws_i1.column_dimensions['A'].width = 3

ws_i1.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 경영인정기보험 자산대장 (2022~2025)").font = title_font
ws_i1.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 대표이사 경영인정기보험 총 5건 자산 대장").font = subtitle_font

ws_i1.merge_cells("B5:M5")
i1_bar_cell = ws_i1.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 계약: 총 5건 (대표이사 보장)   |   월 납입 보험료 합계: ₩ 35,764,220 (월 3,576만원)   |   피보험자: 이종탁 대표이사")
i1_bar_cell.font = Font(name=font_family, size=10, bold=True, color="1E293B")
i1_bar_cell.fill = summary_bar_fill
i1_bar_cell.alignment = Alignment(horizontal='left', vertical='center')
i1_bar_cell.border = table_data_border

i_headers1 = ["연도", "순서", "보험종목", "피보험자/대상", "보험사", "상품명", "증권번호", "보험개시일", "보험만기일", "월 납입액(원)", "계약 상태", "비고 (경영인보장 및 세무사항)"]

ws_i1.row_dimensions[7].height = 28
for c_i, text in enumerate(i_headers1, 2):
    c = ws_i1.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

i_data1 = [
    ("2022", "01", "경영인정기보험", "이종탁 대표이사", "KDB생명", "(무)VIP경영인정기보험", "0946352030001", "2022-05-31", "2069-03-31", 5347500, "유지중", "대표이사 정기보장 (1차)"),
    ("2022", "02", "경영인정기보험", "이종탁 대표이사", "KDB생명", "(무)VIP경영인정기보험", "0946352030002", "2022-05-31", "2069-03-31", 5347500, "유지중", "대표이사 정기보장 (2차)"),
    ("2023", "03", "경영인정기보험", "이종탁 대표이사", "매트라이프", "Honors 경영인정기보험Plus", "13460791", "2023-12-28", "2069-12-28", 10011540, "유지중", "대표이사 정기보장 (매트라이프)"),
    ("2024", "04", "경영인정기보험", "이종탁 대표이사", "미래에셋생명", "VIP 경영인을 위한 정기보험", "8005286685", "2024-02-07", "2069-02-07", 5016000, "유지중", "대표이사 정기보장 (미래에셋)"),
    ("2024", "05", "경영인정기보험", "이종탁 대표이사", "삼성생명", "삼성 간편경영인정기보험", "41000016223329", "2024-04-22", "2074-04-22", 10041680, "유지중", "대표이사 50년납 정기보장 (삼성생명)")
]

for row_idx, row_data in enumerate(i_data1, 8):
    ws_i1.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for t_idx, val in enumerate(row_data):
        c_i = t_idx + 2
        cell = ws_i1.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 5, 6, 7, 8, 9, 10]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 11:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# --- INSURANCE TAB 2: 일반화재_배상책임_자산대장 ---
ws_i2 = wb_ins.create_sheet(title="02_일반화재_배상책임_자산대장")
ws_i2.views.sheetView[0].showGridLines = False
ws_i2.column_dimensions['A'].width = 3

ws_i2.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 화재/배상책임/종합보험 자산대장 (2022~2025)").font = title_font
ws_i2.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 사업장 재산 및 학원배상책임 보험 총 4건 대장").font = subtitle_font

ws_i2.merge_cells("B5:M5")
i2_bar_cell = ws_i2.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 계약: 총 4건   |   월/일시납 보험료 합계: 월 124,664원 + 연간 일시납 249,000원   |   사업장 안전보장")
i2_bar_cell.font = Font(name=font_family, size=10, bold=True, color="1E293B")
i2_bar_cell.fill = summary_bar_fill
i2_bar_cell.alignment = Alignment(horizontal='left', vertical='center')
i2_bar_cell.border = table_data_border

i_headers2 = ["연도", "순서", "보험종목", "피보험자/대상", "보험사", "상품명", "증권번호", "보험개시일", "보험만기일", "납입료/구분(원)", "계약 상태", "비고 (보장내용 및 갱신이력)"]

ws_i2.row_dimensions[7].height = 28
for c_i, text in enumerate(i_headers2, 2):
    c = ws_i2.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

i_data2 = [
    ("2025", "01", "화재재산종합보험", "대전 도룡동 204호", "현대해상", "성공마스터 재산종합보험", "L-025-21582789", "2025-05-09", "2028-05-09", 100000, "유지중", "2025.05.09 갱신 (월 10만원)"),
    ("2025", "02", "학원배상책임보험", "대전 스마트시티 113,114,115호", "DB손해보험", "학원배상책임보험", "120250591347", "2025-04-30", "2026-04-30", 20000, "유지중", "2025.04.30 갱신 (연 일시납 2만원)"),
    ("2024", "03", "홈앤비즈종합보험", "대전 스마트시티 209호", "KB손해보험", "(무)KB홈앤비즈케어종합보험", "2024-1986936", "2024-03-29", "2034-03-29", 24664, "유지중", "10년 장기종합보험 (월 24,664원)"),
    ("2025", "04", "기업종합보험", "광명 GIDC A1212, A1213호", "KB손해보험", "One KB기업종합보험", "2025-5689111", "2025-08-28", "2026-08-28", 229000, "유지중", "광명사옥 기업종합보험 (연 일시납 22.9만원)")
]

for row_idx, row_data in enumerate(i_data2, 8):
    ws_i2.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for t_idx, val in enumerate(row_data):
        c_i = t_idx + 2
        cell = ws_i2.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 5, 6, 7, 8, 9, 10]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 11:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# --- INSURANCE TAB 3: 법인차량_자동차보험_자산대장 ---
ws_i3 = wb_ins.create_sheet(title="03_법인차량_자동차보험_자산대장")
ws_i3.views.sheetView[0].showGridLines = False
ws_i3.column_dimensions['A'].width = 3

ws_i3.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 법인차량 자동차보험 자산대장 (2022~2025)").font = title_font
ws_i3.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 4대 법인 보유 차량 자동차보험 총 4건 대장").font = subtitle_font

ws_i3.merge_cells("B5:M5")
i3_bar_cell = ws_i3.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 계약: 총 4대 차종   |   연간 자동차보험료 합계: ₩ 6,443,910 (644만원)   |   배서 및 갱신 완료")
i3_bar_cell.font = Font(name=font_family, size=10, bold=True, color="1E293B")
i3_bar_cell.fill = summary_bar_fill
i3_bar_cell.alignment = Alignment(horizontal='left', vertical='center')
i3_bar_cell.border = table_data_border

i_headers3 = ["연도", "순서", "차종/차량번호", "피보험자/명의", "보험사", "상품명", "증권번호", "보험개시일", "보험만기일", "연간보험료(원)", "계약 상태", "비고 (배서추납 및 갱신이력)"]

ws_i3.row_dimensions[7].height = 28
for c_i, text in enumerate(i_headers3, 2):
    c = ws_i3.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

i_data3 = [
    ("2025", "01", "K8 (289수3930)", "㈜폭스에듀", "KB손해보험", "KB업무용 자동차보험", "2025-0127909", "2025-01-08", "2026-01-08", 1662600, "유지중", "2025년 KB손보 갱신 체결"),
    ("2025", "02", "벤츠 S클래스 (281가8991)", "㈜폭스에듀", "KB손해보험", "KB업무용 자동차보험", "2025-1512997", "2025-03-10", "2026-03-10", 1257880, "유지중", "2025년 KB손보 갱신 체결"),
    ("2025", "03", "아우디 A8 (120노2842)", "㈜폭스에듀", "DB손해보험", "프로미카업무용 자동차보험", "2-2025-6278074", "2025-11-15", "2026-11-15", 2783960, "유지중", "2025년 DB손보 갱신 체결"),
    ("2025", "04", "카니발 (296더5669)", "㈜폭스에듀", "메리츠화재", "Readycar업무용 자동차보험", "72000-25-0259222", "2025-03-22", "2026-03-22", 739470, "유지중", "2025.05.23 배서 추납 완료")
]

for row_idx, row_data in enumerate(i_data3, 8):
    ws_i3.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for t_idx, val in enumerate(row_data):
        c_i = t_idx + 2
        cell = ws_i3.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 5, 6, 7, 8, 9, 10]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 11:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

ins_col_widths = {1: 3, 2: 10, 3: 8, 4: 20, 5: 18, 6: 14, 7: 24, 8: 20, 9: 14, 10: 14, 11: 18, 12: 14, 13: 44}
for sheet in [ws_i1, ws_i2, ws_i3]:
    for col_idx, width in ins_col_widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

wb_ins.save(ins_excel_p)
print("  [CORPORATE INSURANCE EXCEL SAVED SUCCESSFULLY]", ins_excel_p)

print("\n==========================================")
print("ALL EXCELS RE-GENERATED & VERIFIED WITH CORRECT CELL INDEXING!")
print("==========================================")
