import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FOXCONNECT_ROOT = r"G:\내 드라이브\[FoxConnect]"
REAL_ESTATE_BASE = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\01_부동산_자산관리")
TARGET_DIR = os.path.join(REAL_ESTATE_BASE, "00_연도별_부동산_총괄자산대장")

os.makedirs(TARGET_DIR, exist_ok=True)
excel_path = os.path.join(TARGET_DIR, "[외감_IPO대비]_주식회사_폭스에듀_연도별_부동산_총괄자산대장(2022-2025).xlsx")

print("Updating Excel Master Register with VERIFIED GANGSAN CONTRACT FIGURES...")
print("  - Deposit: 200,000,000 KRW (2억 원)")
print("  - Monthly Rent: 14,700,000 KRW (월 1,470만 원)")

wb = openpyxl.Workbook()
wb.remove(wb.active)

font_family = "맑은 고딕"

title_font = Font(name=font_family, size=15, bold=True, color="0F172A")
subtitle_font = Font(name=font_family, size=9, bold=True, color="475569")

summary_bar_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

data_font = Font(name=font_family, size=10, color="0F172A")
data_bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")

zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

pill_styles = {
    "정상유지": {"fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="166534")},
    "전대차유지": {"fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="1E40AF")},
    "묵시적갱신": {"fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="1E40AF")},
    "만기해지": {"fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="475569")},
    "중도해지": {"fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="991B1B")},
    "소유권보유": {"fill": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), "font": Font(name=font_family, size=9, bold=True, color="6B21A8")}
}

black_border_side = Side(style='thin', color='000000')
black_thick_top_side = Side(style='medium', color='000000')

table_header_border = Border(left=black_border_side, right=black_border_side, top=black_thick_top_side, bottom=black_border_side)
table_data_border = Border(left=black_border_side, right=black_border_side, top=black_border_side, bottom=black_border_side)

# --- 1. TAB 1: 01_임대차_자산대장 ---
ws1 = wb.create_sheet(title="01_임대차_자산대장")
ws1.views.sheetView[0].showGridLines = False
ws1.column_dimensions['A'].width = 3

ws1.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 부동산 임대차 자산대장 (2022~2025)").font = title_font
ws1.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 2025.12.31 기준 유효 임대차 자산 대장").font = subtitle_font

ws1.merge_cells("B5:M5")
# Verified 17 Active Leases Total: Deposit = 1,991,000,000 KRW, Rent = 64,324,500 KRW
bar_cell = ws1.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 자산: 총 17건   |   보증금 잔액 합계: ₩ 1,991,000,000 (19억 9,100만원)   |   월 임대료 합계: ₩ 64,324,500 (VAT 별도)")
bar_cell.font = Font(name=font_family, size=10, bold=True, color="1E293B")
bar_cell.fill = summary_bar_fill
bar_cell.alignment = Alignment(horizontal='left', vertical='center')
bar_cell.border = table_data_border

headers1 = ["연도", "순서", "구분", "물건명 / 주소", "임대인 (명의)", "계약유형", "임대시작일", "임대종료일", "보증금 (원)", "월 임대료 (원)", "계약 상태", "비고 (연도별 변동 및 해지 이력)"]

ws1.row_dimensions[7].height = 28
for c_i, text in enumerate(headers1, 2):
    c = ws1.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

lease_data = [
    ("2022", "01", "임대차", "판교 판교동 612", "박동석(김인숙)", "최초임대차", "2021-07-31", "2023-07-30", 40000000, 3500000, "만기해지", "2023.07.31 계약만기 종료"),
    ("2022", "02", "임대차", "대전 골프존 204호,상담실", "㈜골프존뉴딘홀딩스", "승계계약", "2021-08-17", "2023-08-16", 50000000, 8206000, "정상유지", "2023.09.01 계약연장 및 25.05.31 갱신"),
    ("2022", "03", "임대차", "세종 뱅크빌딩 302호", "타이어뱅크(주)", "최초임대차", "2021-10-01", "2023-09-30", 75000000, 8250000, "만기해지", "2023.10.14 만기 퇴거 완료"),
    ("2022", "04", "임대차", "대전 하우스디어반 B동 721호", "정근호", "최초임대차", "2022-04-07", "2024-04-06", 10000000, 750000, "정상유지", "23년 묵시적갱신 -> 24년/25년 연장계약"),
    ("2022", "05", "임대차", "대전 KCC웰츠타워 1202호", "이선영", "최초임대차", "2022-04-11", "2024-04-10", 20000000, 650000, "묵시적갱신", "2023년~2025년 묵시적 갱신 계속 유지"),
    ("2022", "06", "임대차", "대전 스마트시티 2501호", "김동하", "전세", "2022-06-30", "2024-06-29", 1200000000, 0, "묵시적갱신", "전세보증금 12억원 / 24년, 25년 묵시적갱신"),
    ("2022", "07", "임대차", "광명 GIDC 1214_1215호", "하진우", "최초임대차", "2022-08-01", "2024-07-31", 25000000, 2750000, "묵시적갱신", "24년, 25년 묵시적 갱신 계속 유지"),
    ("2022", "08", "임대차", "대전 월평동577 202호", "류지홍", "최초임대차", "2022-12-12", "2024-12-11", 5000000, 770000, "중도해지", "2023.03.31 중도해지 및 퇴거 완료"),
    ("2023", "09", "임대차", "판교 이레빌딩 3층", "화코스텍인터내셔널㈜", "최초임대차", "2023-01-31", "2025-01-30", 100000000, 11000000, "정상유지", "판교 본점 사옥 (28년 2월까지 5년계약)"),
    ("2023", "10", "임대차", "대전 도룡동 385-28", "김순미", "최초임대차", "2023-02-16", "2025-02-15", 100000000, 6050000, "정상유지", "대전 사옥 리저브 (28년 2월까지 5년계약)"),
    ("2023", "11", "임대차", "광명 센트럴자이 1006호", "김영숙", "최초임대차", "2023-02-24", "2025-02-23", 5000000, 770000, "중도해지", "2023.03.31 중도해지 완료"),
    ("2023", "12", "임대차", "대전 스타빌플러스 511호", "은선희", "최초임대차", "2023-04-08", "2025-04-07", 5000000, 638000, "만기해지", "2025.04.07 만기 계약종료"),
    ("2023", "13", "임대차", "대전 골프존 206호", "㈜골프존뉴딘홀딩스", "승계계약", "2023-08-02", "2025-08-01", 50000000, 5428500, "정상유지", "2023.12.01 법인 승계 완료"),
    ("2023", "14", "임대차", "대전 스마트시티 113호", "강남규", "최초임대차", "2023-10-28", "2025-10-27", 30000000, 1100000, "정상유지", "상가 113호 정상 유지"),
    ("2023", "15", "임대차", "대전 스마트시티 115호", "강남규", "최초임대차", "2023-10-28", "2025-10-27", 30000000, 1100000, "정상유지", "상가 115호 정상 유지"),
    ("2023", "16", "임대차", "서초 강남역리가스퀘어 501호", "이봉순", "최초임대차", "2023-11-10", "2025-11-09", 15000000, 1500000, "만기해지", "2024.11.09 만기 계약종료 (서초본점)"),
    ("2024", "17", "임대차", "대전 스마트시티 209호", "이봉순", "최초임대차", "2024-01-31", "2026-01-30", 30000000, 1500000, "정상유지", "상가 209호 정상 유지"),
    ("2024", "18", "임대차", "대전 하우스디어반 C동 711호", "정근호", "최초임대차", "2024-02-19", "2026-02-18", 10000000, 750000, "중도해지", "2024.02.19 중도해지 처리완료"),
    ("2024", "19", "임대차", "대전 스마트시티 604호", "김동하", "최초임대차", "2024-02-26", "2026-02-25", 50000000, 2500000, "정상유지", "스마트시티 604호 정상 유지"),
    ("2024", "20", "임대차", "가산 대륭포스트타워6차 402_403호", "㈜엠씨에스솔루션", "최초임대차", "2024-02-29", "2026-02-27", 46000000, 4600000, "정상유지", "가산 지식산업센터 402_403호"),
    ("2024", "21", "임대차", "대전 하우스디어반 A동 720호", "정근호", "최초임대차", "2024-04-15", "2026-04-14", 10000000, 790000, "정상유지", "하우스디어반 A동 720호"),
    ("2024", "22", "임대차", "강남 도곡로1길23 지하1층~3층", "박재윤(유한회사 청송)", "최초임대차", "2024-11-01", "2026-10-31", 200000000, 14700000, "정상유지", "강남 사옥 통건물 (보증금 2억/월 1,470만/관리비 219만)"),
    ("2024", "23", "임대차", "대전 갑동 388-1", "데이타이음", "최초임대차", "2024-12-23", "2026-12-22", 20000000, 1200000, "정상유지", "갑동 연구소 및 사업장"),
    ("2025", "24", "임대차", "대전 골프존 104호", "㈜골프존뉴딘홀딩스", "최초임대차", "2025-05-14", "2027-05-13", 10000000, 3148750, "만기해지", "2025.05.31 104호 계약종료분리 (204호는 유지)"),
    ("2025", "25", "임대차", "가산 대륭포스트타워6차 1510호", "㈜엠씨에스솔루션", "최초임대차", "2026-04-29", "2028-04-28", 20000000, 2000000, "정상유지", "가산 1510호 신규확장 체결")
]

for row_idx, row_data in enumerate(lease_data, 8):
    ws1.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for c_i, val in enumerate(row_data, 2):
        cell = ws1.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 6, 7, 8, 9]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 5:
            cell.value = str(val)
            cell.font = data_bold_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif c_i in [10, 11]:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# --- 2. TAB 2: 02_전대차_자산대장 ---
ws2 = wb.create_sheet(title="02_전대차_자산대장")
ws2.views.sheetView[0].showGridLines = False
ws2.column_dimensions['A'].width = 3

ws2.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 부동산 전대차 자산대장 (2022~2025)").font = title_font
ws2.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 2025.12.31 기준 유효 전대차 자산 대장").font = subtitle_font

ws2.merge_cells("B5:M5")
bar_cell2 = ws2.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 전대: 총 7건   |   전대보증금 잔액 합계: ₩ 75,000,000 (7,500만원)   |   월 전대수익 합계: ₩ 8,750,000 (VAT 별도)")
bar_cell2.font = Font(name=font_family, size=10, bold=True, color="1E293B")
bar_cell2.fill = summary_bar_fill
bar_cell2.alignment = Alignment(horizontal='left', vertical='center')
bar_cell2.border = table_data_border

headers2 = ["연도", "순서", "구분", "전대 물건명 / 주소", "전차인 (명의)", "계약유형", "전대시작일", "전대종료일", "전대보증금(원)", "전대월세(원)", "계약 상태", "비고 (전대차 계약이력)"]

ws2.row_dimensions[7].height = 28
for c_i, text in enumerate(headers2, 2):
    c = ws2.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

sublease_data = [
    ("2022", "01", "전대차", "광명 GIDC 1212호 (일부)", "㈜데이타이음", "전대차계약", "2022-03-05", "2023-03-04", 5000000, 500000, "만기해지", "2023.03.04 전대차 계약만기 종료"),
    ("2023", "02", "전대차", "광명 GIDC 1215호 (일부)", "㈜트라이디스", "전대차계약", "2023-06-01", "2024-05-31", 10000000, 1200000, "만기해지", "2024.05.31 전대차 만기 종료"),
    ("2023", "03", "전대차", "광명 GIDC 1214호 (일부)", "㈜데이타이음", "전대차계약", "2023-07-01", "2024-06-30", 15000000, 1750000, "만기해지", "2024.06.30 전대차 만기 종료"),
    ("2023", "04", "전대차", "광명 GIDC 1214호 (일부)", "김두연(DX데이타협회)", "전대차계약", "2023-11-29", "2024-11-28", 10000000, 1000000, "만기해지", "2024.11.28 전대차 만기 종료"),
    ("2024", "05", "전대차", "강남 도곡로1길23 1층", "㈜에스앤에이치트레이딩", "전대차계약", "2024-11-01", "2025-10-31", 20000000, 2000000, "전대차유지", "강남본점 1층 전대차 계약 유지 중"),
    ("2024", "06", "전대차", "강남 도곡로1길23 2층", "㈜실리콘아츠", "전대차계약", "2024-11-01", "2025-10-31", 20000000, 2200000, "전대차유지", "강남본점 2층 전대차 계약 유지 중"),
    ("2025", "07", "전대차", "강남 도곡로1길23 3층", "㈜트라이디스", "전대차계약", "2025-01-24", "2026-01-23", 15000000, 1800000, "전대차유지", "강남본점 3층 일부 전대차 계약 유지"),
    ("2025", "08", "전대차", "광명 GIDC 1212호 (일부)", "㈜엠엘씨", "전대차계약", "2025-06-02", "2026-06-01", 5000000, 650000, "전대차유지", "광명 GIDC 1212호 전대차 갱신 체결"),
    ("2025", "09", "전대차", "가산 대륭포스트타워6차 403호", "정선혜", "전대차계약", "2025-07-01", "2026-06-30", 5000000, 600000, "전대차유지", "가산 403호 전대차 계약 체결"),
    ("2025", "10", "전대차", "강남 도곡로1길23 1층 일부", "한국경찰과학전략센터", "전대차계약", "2025-08-21", "2026-08-20", 5000000, 750000, "전대차유지", "강남 1층 추가 전대차 계약"),
    ("2025", "11", "전대차", "강남 도곡로1길23 1층 일부", "㈜월드유니코어", "전대차계약", "2025-08-21", "2026-08-20", 5000000, 750000, "전대차유지", "강남 1층 추가 전대차 계약")
]

for row_idx, row_data in enumerate(sublease_data, 8):
    ws2.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for c_i, val in enumerate(row_data, 2):
        cell = ws2.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 6, 7, 8, 9]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 5:
            cell.value = str(val)
            cell.font = data_bold_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif c_i in [10, 11]:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# --- 3. TAB 3: 03_소유권_매매_자산대장 ---
ws3 = wb.create_sheet(title="03_소유권_매매_자산대장")
ws3.views.sheetView[0].showGridLines = False
ws3.column_dimensions['A'].width = 3

ws3.cell(row=2, column=2, value="[주식회사 폭스에듀] 연도별 부동산 소유권(매매) 자산대장 (2022~2025)").font = title_font
ws3.cell(row=3, column=2, value="※ 외감/IPO 제출용 (작성 기준일: 2025년 12월 31일) | 2025.12.31 기준 유효 소유권 자산 대장").font = subtitle_font

ws3.merge_cells("B5:M5")
bar_cell3 = ws3.cell(row=5, column=2, value="  [작성 기준일: 2025-12-31 기준]   유지 소유권: 총 2건   |   분양/매매 취득가액 총액: ₩ 1,022,660,000 (10억 2,266만원)   |   보유 형태: 법인 소유 자산")
bar_cell3.font = Font(name=font_family, size=10, bold=True, color="1E293B")
bar_cell3.fill = summary_bar_fill
bar_cell3.alignment = Alignment(horizontal='left', vertical='center')
bar_cell3.border = table_data_border

headers3 = ["연도", "순서", "구분", "소유 자산명 / 주소", "매도인/분양사", "계약유형", "취득일자", "만기일자", "취득가액 (원)", "월 임대료 (원)", "계약 상태", "비고 (자산 취득 및 보유 현황)"]

ws3.row_dimensions[7].height = 28
for c_i, text in enumerate(headers3, 2):
    c = ws3.cell(row=7, column=c_i, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = table_header_border

sale_data = [
    ("2022", "01", "소유권", "광명 GIDC A동 1212호", "하늘기획", "분양권매매", "2022-01-13", "-", 494995000, 0, "소유권보유", "광명 GIDC 자산 소유권 보유 (건물가+토지가)"),
    ("2022", "02", "소유권", "광명 GIDC A동 1213호", "하늘기획", "분양권매매", "2022-01-13", "-", 527665000, 0, "소유권보유", "광명 GIDC 자산 소유권 보유 (건물가+토지가)")
]

for row_idx, row_data in enumerate(sale_data, 8):
    ws3.row_dimensions[row_idx].height = 22
    fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
    
    for c_i, val in enumerate(row_data, 2):
        cell = ws3.cell(row=row_idx, column=c_i)
        cell.border = table_data_border
        cell.fill = fill_to_use

        if c_i in [2, 3]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_bold_font if c_i == 3 else data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i in [4, 6, 7, 8, 9]:
            cell.value = str(val)
            cell.number_format = '@'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_i == 5:
            cell.value = str(val)
            cell.font = data_bold_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif c_i in [10, 11]:
            cell.value = val
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c_i == 12:
            cell.value = str(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if val in pill_styles:
                cell.fill = pill_styles[val]["fill"]
                cell.font = pill_styles[val]["font"]
        elif c_i == 13:
            cell.value = str(val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Auto-fit Column Widths across all 3 sheets
col_widths = {
    1: 3,    # Margin A
    2: 10,   # 연도 B
    3: 8,    # 순서 C
    4: 10,   # 구분 D
    5: 32,   # 물건명 E
    6: 24,   # 명의 F
    7: 14,   # 유형 G
    8: 14,   # 시작 H
    9: 14,   # 종료 I
    10: 18,  # 보증금 J
    11: 16,  # 월세 K
    12: 14,  # 상태 L
    13: 48   # 비고 M
}

for sheet in [ws1, ws2, ws3]:
    for col_idx, width in col_widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(excel_path)
print(f"  [GANGSAN DEPOSIT & RENT UPDATED IN EXCEL SUCCESSFULLY] {excel_path}")
