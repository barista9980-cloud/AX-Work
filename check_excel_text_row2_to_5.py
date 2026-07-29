import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

FOXCONNECT_ROOT = r"G:\내 드라이브\[FoxConnect]"
excel_path = os.path.join(FOXCONNECT_ROOT, r"[총무]업무\01_부동산_자산관리\00_연도별_부동산_총괄자산대장\[외감_IPO대비]_주식회사_폭스에듀_연도별_부동산_총괄자산대장(2022-2025).xlsx")

print("Checking Row 2, Row 3, Row 5 text in Excel:", excel_path)

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} ---")
        for r in range(2, 6):
            row_vals = [str(ws.cell(row=r, column=c).value or '') for c in range(2, 14)]
            non_empty = [v for v in row_vals if v.strip()]
            if non_empty:
                print(f"  Row {r}:", " | ".join(non_empty))
