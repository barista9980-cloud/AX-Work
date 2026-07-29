import os
import sqlite3
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, "data", "real_estate_assets.db")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
GDRIVE_OUTPUT_DIR = r"G:\내 드라이브\[부동산자산] FoxConnect 계약 관리\04_생성_보고서"

os.makedirs(OUTPUT_DIR, exist_ok=True)
if os.path.exists(r"G:\내 드라이브\[부동산자산] FoxConnect 계약 관리"):
    os.makedirs(GDRIVE_OUTPUT_DIR, exist_ok=True)

def generate_annual_report_for_year(year_str):
    target_date = f"{year_str}-12-31" if year_str != "2026_통합" else "2026-12-31"
    
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM real_estate_contracts
            WHERE contract_date <= ?
            ORDER BY contract_date ASC, region ASC, property_info ASC
        """, (target_date,))
        rows = [dict(r) for r in cursor.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year_str}_부동산자산현황"

    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="맑은 고딕", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    even_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    headers = [
        "연번", "기준일자", "계약체결일", "권역", "물건지 및 호수", "차수",
        "계약유형", "당사 구분", "임대인(전대인)", "임차인(전차인)", "파일명"
    ]
    
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for idx, r in enumerate(rows, 1):
        row_data = [
            idx,
            target_date,
            r["contract_date"],
            r["region"],
            r["property_info"],
            r["sequence"],
            r["contract_type"],
            r["our_role"],
            r["party_a_normalized"],
            r["party_b_normalized"],
            r["filename"]
        ]
        ws.append(row_data)
        
        row_num = idx + 1
        is_even = (idx % 2 == 0)
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = even_row_fill
                
            if col_num in [1, 2, 3, 4, 6, 7, 8]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            korean_count = sum(1 for c in val if ord(c) > 127)
            val_len = len(val) + korean_count
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename_base = f"{year_str}년12월31일기준_부동산자산현황_보고서" if year_str != "2026_통합" else "FoxConnect_부동산_자산통합대장_2026최신"
    
    # Save Excel & CSV
    xlsx_local = os.path.join(OUTPUT_DIR, f"{filename_base}.xlsx")
    xlsx_gdrive = os.path.join(GDRIVE_OUTPUT_DIR, f"{filename_base}.xlsx")
    wb.save(xlsx_local)
    if os.path.exists(GDRIVE_OUTPUT_DIR):
        wb.save(xlsx_gdrive)

    csv_local = os.path.join(OUTPUT_DIR, f"{filename_base}.csv")
    csv_gdrive = os.path.join(GDRIVE_OUTPUT_DIR, f"{filename_base}.csv")
    
    for p in [csv_local, csv_gdrive]:
        if os.path.dirname(p) and os.path.exists(os.path.dirname(p)):
            with open(p, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for idx, r in enumerate(rows, 1):
                    writer.writerow([
                        idx,
                        target_date,
                        r["contract_date"],
                        r["region"],
                        r["property_info"],
                        r["sequence"],
                        r["contract_type"],
                        r["our_role"],
                        r["party_a_normalized"],
                        r["party_b_normalized"],
                        r["filename"]
                    ])
                    
    print(f"Generated report for {year_str}: {len(rows)} records.")

def main():
    years = ["2021", "2022", "2023", "2024", "2025", "2026_통합"]
    for y in years:
        generate_annual_report_for_year(y)

if __name__ == "__main__":
    main()
